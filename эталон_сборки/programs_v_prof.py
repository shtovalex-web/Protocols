# -*- coding: utf-8 -*-
"""
Лист V_PROF в Programs_base.xlsx: поддержка старого и нового формата.

Старый: A — профессия; B/ПП/СИЗ — полный текст; столбцы 5+ — тексты программ «В».
Новый: A — профессия; B — ПП, C — СИЗ, D — «Б»; столбцы с номерами 6,7,… — «Да»;
       шапка протокола «В» — лист V столбец C; таблица и Минтруд «В» — лист V столбец B.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any, Literal

from v_program_registry_match import (
    header_title_for_v_program_id,
    norm_profession_key,
    table_text_for_v_prof_fragment,
    table_text_for_v_program_id,
)

V_PROF_SHEET_NAME = "V_PROF"

_YES_MARKERS = frozenset(
    {
        "да",
        "yes",
        "y",
        "1",
        "+",
        "x",
        "х",
        "true",
        "истина",
    }
)


@dataclass(frozen=True)
class VProfLayout:
    format: Literal["legacy", "matrix"]
    col_profession: int  # 0-based
    col_b: int
    col_pp: int
    col_siz: int
    v_marker_columns: tuple[tuple[int, int], ...]  # (0-based col, program_id)
    last_col: int

    @property
    def col_b_one_based(self) -> int:
        return self.col_b + 1

    @property
    def col_pp_one_based(self) -> int:
        return self.col_pp + 1

    @property
    def col_siz_one_based(self) -> int:
        return self.col_siz + 1


@dataclass(frozen=True)
class VProfProfessionCandidate:
    """Строка V_PROF (столбец A) с оценкой близости к введённой должности."""

    profession: str
    score: int
    v_program_count: int


@dataclass(frozen=True)
class VProfProfessionMatch:
    """Результат сопоставления введённой должности со строкой V_PROF."""

    profession_input: str
    profession_matched: str
    score: int  # 0..3 как в _select_best_row_by_profession_col_a
    v_program_count: int
    layout_format: str


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", s)


def _parse_program_id_from_header(header: str) -> int | None:
    if not header:
        return None
    m = re.match(r"^(\d+)\b", header)
    if m:
        return int(m.group(1))
    if header.isdigit():
        return int(header)
    return None


def _is_yes_marker(value: object) -> bool:
    if value is None:
        return False
    s = str(value).strip().lower().replace("ё", "е")
    if not s:
        return False
    if s in _YES_MARKERS:
        return True
    return s in ("✓", "v", "√")


def _detect_layout_from_header(header: tuple[Any, ...]) -> VProfLayout:
    cols = len(header)
    col_pp = 1
    col_siz = 2
    col_b = 3
    fmt: Literal["legacy", "matrix"] = "legacy"
    v_markers: list[tuple[int, int]] = []

    h1 = _normalize_header(header[1]) if cols > 1 else ""
    h2 = _normalize_header(header[2]) if cols > 2 else ""
    h3 = _normalize_header(header[3]) if cols > 3 else ""

    if ("пп" in h1 or "первой помощ" in h1) and (
        "сиз" in h2 or "средств" in h2 and "защит" in h2
    ):
        fmt = "matrix"
        col_pp, col_siz, col_b = 1, 2, 3
    elif h1 in ("б", "b") or "программа б" in h1 or h1.endswith(" б"):
        fmt = "legacy"
        col_b, col_pp, col_siz = 1, 2, 3
    else:
        for j in range(1, cols):
            hn = _normalize_header(header[j])
            if not hn:
                continue
            if "пп" in hn and "программ" in hn:
                col_pp = j
            elif "сиз" in hn and "программ" in hn:
                col_siz = j
            elif re.search(r"(^|\s)б(\s|$)|программ.*\bб\b", hn):
                col_b = j

    for j in range(4, cols):
        pid = _parse_program_id_from_header(_normalize_header(header[j]))
        if pid is not None:
            v_markers.append((j, pid))
            fmt = "matrix"

    if not v_markers and cols > 4:
        fmt = "legacy"
        for j in range(4, cols):
            v_markers.append((j, j - 3))

    last_col = max((c for c, _ in v_markers), default=cols - 1)
    last_col = max(last_col, col_b, col_pp, col_siz)

    return VProfLayout(
        format=fmt,
        col_profession=0,
        col_b=col_b,
        col_pp=col_pp,
        col_siz=col_siz,
        v_marker_columns=tuple(v_markers),
        last_col=last_col,
    )


def _read_header_row(path: Path) -> tuple[Any, ...]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        names = {n.lower(): n for n in wb.sheetnames}
        sn = names.get(V_PROF_SHEET_NAME.lower())
        if not sn:
            return tuple()
        ws = wb[sn]
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        return tuple(row) if row else tuple()
    finally:
        wb.close()


@lru_cache(maxsize=8)
def get_v_prof_layout(path_str: str, mtime_ns: int) -> VProfLayout:
    """Кэш разметки V_PROF по пути и mtime файла."""
    _ = mtime_ns
    header = _read_header_row(Path(path_str))
    if not header:
        return VProfLayout(
            format="legacy",
            col_profession=0,
            col_b=1,
            col_pp=2,
            col_siz=3,
            v_marker_columns=tuple((j, j - 3) for j in range(4, 22)),
            last_col=21,
        )
    return _detect_layout_from_header(header)


def v_prof_layout_for_path(path: Path) -> VProfLayout:
    p = path.resolve()
    mtime = p.stat().st_mtime_ns if p.is_file() else 0
    return get_v_prof_layout(str(p), mtime)


def v_prof_row_program_header_titles(
    row: tuple[Any, ...],
    layout: VProfLayout,
    v_registry_rows: list[tuple[str, str, str, float | None]],
) -> list[str]:
    """Шапка протокола / FG: «В» — столбец C листа V (матрица «Да» + ID в шапке V_PROF)."""
    if layout.format == "legacy":
        from v_program_registry_match import match_v_registry_fragment

        out: list[str] = []
        for idx, _ in layout.v_marker_columns:
            if idx >= len(row):
                break
            cell = row[idx]
            if cell is None:
                continue
            t = str(cell).strip()
            if t and not _is_yes_marker(t):
                m = match_v_registry_fragment(t, v_registry_rows)
                if m and (m[0] or "").strip():
                    out.append((m[0] or "").strip())
                else:
                    out.append(t)
        return out

    titles: list[str] = []
    for col_idx, prog_id in layout.v_marker_columns:
        if col_idx >= len(row):
            continue
        if not _is_yes_marker(row[col_idx]):
            continue
        title = header_title_for_v_program_id(prog_id, v_registry_rows)
        titles.append(title if title else str(prog_id))
    return titles


def v_prof_row_program_table_fragments(
    row: tuple[Any, ...],
    layout: VProfLayout,
    v_registry_rows: list[tuple[str, str, str, float | None]],
) -> list[str]:
    """Таблица протокола / Минтруд «В»: столбец B листа V (матрица «Да» + ID в шапке V_PROF)."""
    if layout.format == "legacy":
        out: list[str] = []
        for idx, _ in layout.v_marker_columns:
            if idx >= len(row):
                break
            cell = row[idx]
            if cell is None:
                continue
            t = str(cell).strip()
            if t and not _is_yes_marker(t):
                out.append(table_text_for_v_prof_fragment(t, v_registry_rows))
        return out

    fragments: list[str] = []
    for col_idx, prog_id in layout.v_marker_columns:
        if col_idx >= len(row):
            continue
        if not _is_yes_marker(row[col_idx]):
            continue
        text = table_text_for_v_program_id(prog_id, v_registry_rows)
        fragments.append(text if text else str(prog_id))
    return fragments


def v_prof_row_program_titles(
    row: tuple[Any, ...],
    layout: VProfLayout,
    v_registry_rows: list[tuple[str, str, str, float | None]],
) -> list[str]:
    """Совместимость: заголовки для шапки (столбец C листа V)."""
    return v_prof_row_program_header_titles(row, layout, v_registry_rows)


def count_v_programs_for_row(
    row: tuple[Any, ...] | None, layout: VProfLayout
) -> int:
    if not row:
        return 0
    if layout.format == "legacy":
        n = 0
        for idx, _ in layout.v_marker_columns:
            if idx < len(row) and row[idx] is not None and str(row[idx]).strip():
                if not _is_yes_marker(row[idx]):
                    n += 1
        return n
    return sum(
        1
        for idx, _ in layout.v_marker_columns
        if idx < len(row) and _is_yes_marker(row[idx])
    )


_V_PROF_SKIP_HEADERS = frozenset(
    {
        "профессия",
        "должность",
        "специальность",
        "фио",
        "п/п",
        "наименование",
    }
)


@lru_cache(maxsize=8)
def _load_v_prof_profession_rows(
    path_str: str, mtime_ns: int
) -> tuple[tuple[str, tuple[Any, ...]], ...]:
    """Все строки V_PROF: (профессия из A, полная строка)."""
    _ = mtime_ns
    from openpyxl import load_workbook

    path = Path(path_str)
    layout = v_prof_layout_for_path(path)
    max_col = layout.last_col + 1
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        names = {n.lower(): n for n in wb.sheetnames}
        sn = names.get(V_PROF_SHEET_NAME.lower())
        if not sn:
            return tuple()
        ws = wb[sn]
        out: list[tuple[str, tuple[Any, ...]]] = []
        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row or 500, 2000),
            min_col=1,
            max_col=max_col,
            values_only=True,
        ):
            if not row or row[0] is None:
                continue
            prof = str(row[0]).strip()
            if not prof:
                continue
            key = norm_profession_key(prof)
            if not key or key in _V_PROF_SKIP_HEADERS:
                continue
            out.append((prof, tuple(row)))
        return tuple(out)
    finally:
        wb.close()


def similar_professions_in_v_prof(
    path: Path,
    profession: str,
    *,
    limit: int = 5,
    min_score: int = 1,
) -> list[VProfProfessionCandidate]:
    """Топ похожих профессий из столбца A листа V_PROF для подсказок в интерфейсе."""
    pr = (profession or "").strip()
    if not pr or not path.is_file():
        return []
    p = path.resolve()
    mtime = p.stat().st_mtime_ns
    rows = _load_v_prof_profession_rows(str(p), mtime)
    if not rows:
        return []
    layout = v_prof_layout_for_path(p)
    target = norm_profession_key(pr)
    scored: list[VProfProfessionCandidate] = []
    seen: set[str] = set()
    for prof, row in rows:
        key = norm_profession_key(prof)
        if key in seen:
            continue
        score = 0
        if key == target:
            score = 3
        elif target in key or key in target:
            score = 2
        elif min_score <= 1:
            # Частичное совпадение по словам (длинные должности)
            tw = set(target.split())
            kw = set(key.split())
            if tw and kw and (tw <= kw or kw <= tw or len(tw & kw) >= 2):
                score = 1
        if score < min_score:
            continue
        seen.add(key)
        scored.append(
            VProfProfessionCandidate(
                profession=prof,
                score=score,
                v_program_count=count_v_programs_for_row(row, layout),
            )
        )
    scored.sort(
        key=lambda c: (-c.score, -c.v_program_count, c.profession.lower())
    )
    return scored[: max(1, limit)]


def match_profession_in_v_prof(
    path: Path, profession: str, *, select_row_fn
) -> VProfProfessionMatch | None:
    """Сопоставление должности из rabotnik с колонкой «Профессия» листа V_PROF."""
    pr = (profession or "").strip()
    if not pr or not path.is_file():
        return None
    row = select_row_fn(path, pr)
    if not row or not row[0]:
        return None
    matched = str(row[0]).strip()
    target = norm_profession_key(pr)
    matched_n = norm_profession_key(matched)
    score = 0
    if matched_n == target:
        score = 3
    elif target in matched_n or matched_n in target:
        score = 2
    else:
        score = 1
    layout = v_prof_layout_for_path(path)
    return VProfProfessionMatch(
        profession_input=pr,
        profession_matched=matched,
        score=score,
        v_program_count=count_v_programs_for_row(row, layout),
        layout_format=layout.format,
    )
