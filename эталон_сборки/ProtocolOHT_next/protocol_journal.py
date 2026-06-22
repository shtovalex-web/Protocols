# -*- coding: utf-8 -*-
"""Журнал сохранённых протоколов в SQLite (таблица protocols)."""

from __future__ import annotations

import csv
import json
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from employees_io import EmployeeRecord
from protocol_paths import database_path

PROTOCOL_JOURNAL_KIND_OT = "OT"
PROTOCOL_JOURNAL_KIND_TECH = "TECH"


def employee_record_to_meta_dict(emp: EmployeeRecord) -> dict[str, str]:
    """Снимок строки сотрудника для export_meta_json (должности на момент формирования протокола)."""
    return {
        "fio": (emp.fio or "").strip(),
        "profession": (emp.profession or "").strip(),
        "subdivision": (emp.subdivision or "").strip(),
        "profession2": (emp.profession2 or "").strip(),
        "snils": (emp.snils or "").strip(),
    }


def build_protocol_export_meta_json(
    program_keys: list[str],
    program_titles: list[str],
    *,
    protocol_no_formatted: str,
    persons_raw: list[EmployeeRecord] | None = None,
    persons_row_source: list[EmployeeRecord] | None = None,
    face_sheet_profession: str | None = None,
    v_prof_enabled_by_fio: dict[str, frozenset[str]] | None = None,
    v_prof_main_by_fio: dict[str, str] | None = None,
) -> str:
    """JSON журнала для Минтруда: программы, номер протокола, сотрудники с должностями."""
    titles_map = {str(k): str(t) for k, t in zip(program_keys, program_titles)}
    payload: dict[str, Any] = {
        "program_keys": [str(k) for k in program_keys],
        "program_titles": titles_map,
        "protocol_no": (protocol_no_formatted or "").strip(),
    }
    if persons_raw:
        payload["persons_raw"] = [
            employee_record_to_meta_dict(p) for p in persons_raw if (p.fio or "").strip()
        ]
    if persons_row_source:
        payload["persons_row_source"] = [
            employee_record_to_meta_dict(p)
            for p in persons_row_source
            if (p.fio or "").strip()
        ]
    fs = (face_sheet_profession or "").strip()
    if fs:
        payload["face_sheet_profession"] = fs
    if v_prof_enabled_by_fio:
        payload["v_prof_enabled_by_fio"] = {
            str(k): sorted(str(x) for x in v)
            for k, v in v_prof_enabled_by_fio.items()
        }
    if v_prof_main_by_fio:
        payload["v_prof_main_by_fio"] = {
            str(k): str(v) for k, v in v_prof_main_by_fio.items() if str(v).strip()
        }
    return json.dumps(payload, ensure_ascii=False)


def save_protocol(
    fio: str,
    topic: str,
    date: str,
    grade: str,
    content: str,
    export_meta_json: str | None = None,
    *,
    protocol_kind: str = PROTOCOL_JOURNAL_KIND_OT,
) -> int:
    """Сохранить запись журнала; при том же №/дате/виде — обновить, не дублировать."""
    kind = (protocol_kind or PROTOCOL_JOURNAL_KIND_OT).strip() or PROTOCOL_JOURNAL_KIND_OT
    with sqlite3.connect(database_path()) as conn:
        existing_id = _find_protocol_journal_id_for_update(
            conn,
            protocol_kind=kind,
            date=date,
            export_meta_json=export_meta_json,
            fio=fio,
        )
        if existing_id is not None:
            conn.execute(
                """
                UPDATE protocols
                SET fio = ?, topic = ?, date = ?, grade = ?, content = ?,
                    export_meta_json = ?, protocol_kind = ?,
                    created_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (fio, topic, date, grade, content, export_meta_json, kind, existing_id),
            )
            conn.commit()
            return existing_id
        cur = conn.execute(
            """
            INSERT INTO protocols (fio, topic, date, grade, content, export_meta_json, protocol_kind)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (fio, topic, date, grade, content, export_meta_json, kind),
        )
        conn.commit()
        return int(cur.lastrowid)


def _find_protocol_journal_id_for_update(
    conn: sqlite3.Connection,
    *,
    protocol_kind: str,
    date: str,
    export_meta_json: str | None,
    fio: str,
) -> int | None:
    """Ключ перезаписи: вид + дата + № протокола из meta; иначе вид + дата + ФИО."""
    kind = (protocol_kind or PROTOCOL_JOURNAL_KIND_OT).strip() or PROTOCOL_JOURNAL_KIND_OT
    date_s = (date or "").strip()
    pn = export_meta_protocol_no(export_meta_json)
    if kind == PROTOCOL_JOURNAL_KIND_TECH:
        cur = conn.execute(
            """
            SELECT id, fio, export_meta_json FROM protocols
            WHERE date = ? AND protocol_kind = ?
            """,
            (date_s, PROTOCOL_JOURNAL_KIND_TECH),
        )
    else:
        cur = conn.execute(
            """
            SELECT id, fio, export_meta_json FROM protocols
            WHERE date = ? AND COALESCE(protocol_kind, ?) = ?
            """,
            (date_s, PROTOCOL_JOURNAL_KIND_OT, PROTOCOL_JOURNAL_KIND_OT),
        )
    rows = cur.fetchall()
    if pn:
        for rid, _fio, meta in rows:
            if export_meta_protocol_no(meta) == pn:
                return int(rid)
    nf = _norm_fio_journal_key(fio)
    if nf:
        for rid, row_fio, _meta in rows:
            if _norm_fio_journal_key(row_fio or "") == nf:
                return int(rid)
    return None


def _journal_row_dedupe_key(record: dict[str, Any]) -> tuple[str, str, str]:
    """Ключ уникальности записи журнала (как при upsert в save_protocol)."""
    kind = (record.get("protocol_kind") or PROTOCOL_JOURNAL_KIND_OT).strip() or PROTOCOL_JOURNAL_KIND_OT
    date_s = (record.get("date") or "").strip()
    pn = export_meta_protocol_no(record.get("export_meta_json"))
    if pn:
        return (kind, date_s, f"no:{pn.casefold()}")
    nf = _norm_fio_journal_key(record.get("fio") or "")
    return (kind, date_s, f"fio:{nf}")


def dedupe_journal_records_for_export(
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Одна запись журнала на протокол (оставляем с наибольшим id)."""
    best: dict[tuple[str, str, str], dict[str, Any]] = {}
    for r in records:
        key = _journal_row_dedupe_key(r)
        prev = best.get(key)
        if prev is None or int(r.get("id") or 0) > int(prev.get("id") or 0):
            best[key] = r
    if not best:
        return list(records)
    out = list(best.values())
    out.sort(key=lambda row: int(row.get("id") or 0), reverse=True)
    return out


def get_protocols_journal_display(protocol_kind: str | None = None) -> list[dict[str, Any]]:
    """Записи журнала для списков в интерфейсе — без дублей (актуальная версия протокола)."""
    return dedupe_journal_records_for_export(get_all_protocols(protocol_kind))


def purge_duplicate_protocol_journal_rows(db_path: Path | None = None) -> int:
    """Удалить из БД старые дубликаты журнала, оставив запись с наибольшим id."""
    path = db_path if db_path is not None else database_path()
    with sqlite3.connect(path) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.execute(
            "SELECT id, fio, date, export_meta_json, protocol_kind FROM protocols"
        )
        rows = [dict(row) for row in cur.fetchall()]
        keep_ids: set[int] = set()
        best_by_key: dict[tuple[str, str, str], int] = {}
        for r in rows:
            rid = int(r["id"])
            key = _journal_row_dedupe_key(r)
            prev = best_by_key.get(key)
            if prev is None or rid > prev:
                best_by_key[key] = rid
        keep_ids.update(best_by_key.values())
        delete_ids = [int(r["id"]) for r in rows if int(r["id"]) not in keep_ids]
        if not delete_ids:
            return 0
        conn.executemany("DELETE FROM protocols WHERE id = ?", [(i,) for i in delete_ids])
        conn.commit()
        return len(delete_ids)


def get_all_protocols(protocol_kind: str | None = None) -> list[dict[str, Any]]:
    with sqlite3.connect(database_path()) as conn:
        conn.row_factory = sqlite3.Row
        if protocol_kind == PROTOCOL_JOURNAL_KIND_TECH:
            cur = conn.execute(
                """
                SELECT id, fio, topic, date, grade, content, export_meta_json, created_at, protocol_kind
                FROM protocols
                WHERE protocol_kind = ?
                ORDER BY id DESC
                """,
                (PROTOCOL_JOURNAL_KIND_TECH,),
            )
        elif protocol_kind == PROTOCOL_JOURNAL_KIND_OT:
            cur = conn.execute(
                """
                SELECT id, fio, topic, date, grade, content, export_meta_json, created_at, protocol_kind
                FROM protocols
                WHERE COALESCE(protocol_kind, ?) = ?
                ORDER BY id DESC
                """,
                (PROTOCOL_JOURNAL_KIND_OT, PROTOCOL_JOURNAL_KIND_OT),
            )
        else:
            cur = conn.execute(
                """
                SELECT id, fio, topic, date, grade, content, export_meta_json, created_at, protocol_kind
                FROM protocols
                ORDER BY id DESC
                """
            )
        return [dict(row) for row in cur.fetchall()]


def clear_protocol_journal(protocol_kind: str | None = None) -> int:
    """
    Удаляет строки журнала protocols.
    protocol_kind None — все записи; OT — только охрана труда (и старые без вида); TECH — только тех. протоколы.
    """
    with sqlite3.connect(database_path()) as conn:
        if protocol_kind == PROTOCOL_JOURNAL_KIND_TECH:
            conn.execute(
                "DELETE FROM protocols WHERE protocol_kind = ?",
                (PROTOCOL_JOURNAL_KIND_TECH,),
            )
        elif protocol_kind == PROTOCOL_JOURNAL_KIND_OT:
            conn.execute(
                "DELETE FROM protocols WHERE COALESCE(protocol_kind, ?) = ?",
                (PROTOCOL_JOURNAL_KIND_OT, PROTOCOL_JOURNAL_KIND_OT),
            )
        else:
            conn.execute("DELETE FROM protocols")
        ch = conn.execute("SELECT changes()").fetchone()
        conn.commit()
        return int(ch[0]) if ch and ch[0] is not None else 0


def _norm_fio_journal_key(fio: str) -> str:
    t = (fio or "").strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", t)


def export_meta_protocol_no(meta_json: str | None) -> str:
    """Номер протокола (N-М-ГГ), сохранённый при формировании в export_meta_json."""
    if not meta_json:
        return ""
    try:
        d = json.loads(meta_json)
        v = d.get("protocol_no")
        return (v if isinstance(v, str) else str(v or "")).strip()
    except (json.JSONDecodeError, TypeError, AttributeError):
        return ""


def journal_kind_label(protocol_kind: str | None) -> str:
    k = (protocol_kind or PROTOCOL_JOURNAL_KIND_OT).strip()
    if k == PROTOCOL_JOURNAL_KIND_TECH:
        return "Технический"
    return "Охрана труда"


def format_journal_list_line(r: dict[str, Any]) -> str:
    """Строка списка в окне журнала: № протокола, дата, оценка, ФИО, тема, id, время записи."""
    pn = export_meta_protocol_no(r.get("export_meta_json")) or "—"
    tid = r.get("id", "")
    dt = (r.get("date") or "").strip()
    gr = (r.get("grade") or "").strip()
    fio = (r.get("fio") or "").strip()
    if len(fio) > 48:
        fio = fio[:45] + "…"
    top = (r.get("topic") or "").strip()
    if len(top) > 32:
        top = top[:29] + "…"
    ca = (r.get("created_at") or "").strip()
    return f"№ {pn:<14}  {dt}  {gr}  |  {fio}  |  {top}  |  id={tid}  {ca}"


def _split_multi_fio_text(fio_field: str) -> list[str]:
    """Несколько ФИО в одной строке журнала (через запятую или «;»)."""
    t = (fio_field or "").strip()
    if not t:
        return []
    if "," not in t and ";" not in t:
        return [t]
    parts = [p.strip() for p in re.split(r"[,;]", t) if p.strip()]
    return parts if parts else [t]


def person_fios_from_journal_row(r: dict[str, Any]) -> list[str]:
    """
    Список ФИО по записи журнала: из export_meta_json (persons_raw) или разбор поля fio.
    """
    meta = r.get("export_meta_json")
    if meta:
        try:
            data = json.loads(meta)
            raw = data.get("persons_raw")
            if isinstance(raw, list) and raw:
                out: list[str] = []
                seen: set[str] = set()
                for item in raw:
                    if not isinstance(item, dict):
                        continue
                    f = (item.get("fio") or "").strip()
                    if not f:
                        continue
                    k = _norm_fio_journal_key(f)
                    if k in seen:
                        continue
                    seen.add(k)
                    out.append(f)
                if out:
                    return out
        except (json.JSONDecodeError, TypeError, AttributeError):
            pass
    single = (r.get("fio") or "").strip()
    if not single:
        return []
    split = _split_multi_fio_text(single)
    deduped: list[str] = []
    seen2: set[str] = set()
    for f in split:
        k = _norm_fio_journal_key(f)
        if k not in seen2:
            seen2.add(k)
            deduped.append(f)
    return deduped


def expand_journal_rows_for_registry_export(
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    Запись журнала на нескольких человек — по одной строке реестра на каждое ФИО
    (охрана труда и тех. протоколы).
    """
    expanded: list[dict[str, Any]] = []
    for r in rows:
        fios = person_fios_from_journal_row(r)
        if len(fios) <= 1:
            expanded.append(r)
            continue
        for fio in fios:
            row_copy = dict(r)
            row_copy["fio"] = fio
            expanded.append(row_copy)
    return expanded


def journal_row_registry_fields(r: dict[str, Any]) -> dict[str, str]:
    """Поля одной записи для выгрузки реестра в Excel/CSV."""
    return {
        "protocol_no": export_meta_protocol_no(r.get("export_meta_json")),
        "date": (r.get("date") or "").strip(),
        "grade": (r.get("grade") or "").strip(),
        "fio": (r.get("fio") or "").strip(),
        "topic": (r.get("topic") or "").strip(),
        "journal_kind": journal_kind_label(r.get("protocol_kind")),
        "created_at": (r.get("created_at") or "").strip(),
        "id": str(r.get("id") or ""),
    }


_REGISTRY_COLUMNS: tuple[tuple[str, str], ...] = (
    ("protocol_no", "№ протокола"),
    ("date", "Дата"),
    ("grade", "Оценка"),
    ("fio", "ФИО"),
    ("topic", "Программы / тема"),
    ("journal_kind", "Вид"),
    ("created_at", "Запись в журнале"),
    ("id", "ID в базе"),
)


def default_journal_registry_export_path(
    *,
    journal_kind: str = PROTOCOL_JOURNAL_KIND_OT,
) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    suffix = "тех" if journal_kind == PROTOCOL_JOURNAL_KIND_TECH else "ОТ"
    return Path.cwd() / f"Реестр_протоколов_{suffix}_{stamp}.xlsx"


def export_protocol_journal_registry(
    path: Path,
    rows: list[dict[str, Any]],
) -> int:
    """
    Выгрузка реестра сформированных протоколов в .xlsx (или .csv при отсутствии openpyxl).
    Возвращает число строк в файле (после разбивки записей на несколько ФИО).
    """
    path = Path(path)
    export_rows = expand_journal_rows_for_registry_export(rows)
    fields_rows = [journal_row_registry_fields(r) for r in export_rows]
    keys = [k for k, _ in _REGISTRY_COLUMNS]
    headers = [h for _, h in _REGISTRY_COLUMNS]

    if path.suffix.lower() == ".csv":
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(headers)
            for fr in fields_rows:
                w.writerow([fr.get(k, "") for k in keys])
        return len(fields_rows)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        csv_path = path.with_suffix(".csv")
        export_protocol_journal_registry(csv_path, rows)
        if path.suffix.lower() != ".csv":
            raise RuntimeError(
                "Для Excel нужен пакет openpyxl (pip install openpyxl). "
                f"Реестр сохранён как CSV: {csv_path}"
            ) from None
        return

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр"
    bold = Font(bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold
    for row_idx, fr in enumerate(fields_rows, start=2):
        for col, key in enumerate(keys, start=1):
            ws.cell(row=row_idx, column=col, value=fr.get(key, ""))
    for col in range(1, len(headers) + 1):
        letter = ws.cell(row=1, column=col).column_letter
        max_len = len(str(headers[col - 1]))
        for row_idx in range(2, min(2 + len(fields_rows), 202)):
            v = ws.cell(row=row_idx, column=col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)
    wb.save(path)
    return len(fields_rows)


def _export_meta_protocol_no_value(meta_json: str | None) -> str:
    return export_meta_protocol_no(meta_json)


def fetch_protocol_journal_rows_by_date(
    date_str: str, *, journal_kind: str = PROTOCOL_JOURNAL_KIND_OT
) -> list[dict[str, Any]]:
    ds = (date_str or "").strip()
    kind = (journal_kind or PROTOCOL_JOURNAL_KIND_OT).strip() or PROTOCOL_JOURNAL_KIND_OT
    with sqlite3.connect(database_path()) as conn:
        conn.row_factory = sqlite3.Row
        if kind == PROTOCOL_JOURNAL_KIND_TECH:
            cur = conn.execute(
                "SELECT id, fio, export_meta_json FROM protocols WHERE date = ? AND protocol_kind = ?",
                (ds, PROTOCOL_JOURNAL_KIND_TECH),
            )
        else:
            cur = conn.execute(
                """
                SELECT id, fio, export_meta_json FROM protocols
                WHERE date = ? AND COALESCE(protocol_kind, ?) = ?
                """,
                (ds, PROTOCOL_JOURNAL_KIND_OT, PROTOCOL_JOURNAL_KIND_OT),
            )
        return [dict(row) for row in cur.fetchall()]


def delete_protocol_journal_rows_by_ids(ids: list[int]) -> int:
    if not ids:
        return 0
    with sqlite3.connect(database_path()) as conn:
        # Плейсхолдеры только «?»; сами id передаются отдельно (не SQL-инъекция).
        ph = ",".join("?" * len(ids))
        conn.execute("DELETE FROM protocols WHERE id IN (" + ph + ")", ids)  # nosec B608
        ch = conn.execute("SELECT changes()").fetchone()
        conn.commit()
        return int(ch[0]) if ch and ch[0] is not None else 0


def journal_ids_and_error_for_per_employee_batch(
    date_str: str,
    fio_and_pn_fmt: list[tuple[str, str]],
    *,
    journal_kind: str = PROTOCOL_JOURNAL_KIND_OT,
) -> tuple[list[int], str | None, list[str]]:
    """
    По дате и планируемой партии (ФИО, полный № как в export_meta_json):
    строки журнала с этими номерами за дату — для предупреждения о перезаписи (upsert).
    err — только если в одной партии одному номеру соответствуют разные ФИО.
    notes — краткие предупреждения для окна подтверждения перезаписи.
    """
    pn_to_fio_norm: dict[str, str] = {}
    planned_pns: set[str] = set()
    for fio, pn_fmt in fio_and_pn_fmt:
        pn = (pn_fmt or "").strip()
        if not pn:
            continue
        planned_pns.add(pn)
        nf = _norm_fio_journal_key(fio)
        prev = pn_to_fio_norm.get(pn)
        if prev is not None and prev != nf:
            return (
                [],
                "Внутренняя ошибка: дублируется номер протокола в одной партии.",
                [],
            )
        pn_to_fio_norm[pn] = nf

    rows = fetch_protocol_journal_rows_by_date(date_str, journal_kind=journal_kind)
    ids_del: list[int] = []
    notes: list[str] = []
    seen_id: set[int] = set()
    for row in rows:
        pn = _export_meta_protocol_no_value(row.get("export_meta_json"))
        if not pn or pn not in planned_pns:
            continue
        rid = int(row["id"])
        if rid in seen_id:
            continue
        seen_id.add(rid)
        ids_del.append(rid)
        jf = _norm_fio_journal_key(row.get("fio") or "")
        if pn in pn_to_fio_norm and jf != pn_to_fio_norm[pn]:
            raw_fio = (row.get("fio") or "").strip()
            notes.append(
                f"№ {pn}: в журнале было ФИО «{raw_fio}»; в партии этот номер — у другого человека."
            )
    return ids_del, None, notes
