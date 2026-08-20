# -*- coding: utf-8 -*-
"""Чтение сотрудников и кандидатов в комиссию из Excel (Data_base.xlsx)."""

from __future__ import annotations

import logging
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

_logger = logging.getLogger(__name__)


class EmployeeExcelError(Exception):
    """Ошибка чтения списка сотрудников из Excel."""


def _workbook_path_for_openpyxl(path: Path) -> Path:
    from bundle_integration import BundleOfficeConvertError, resolve_openpyxl_workbook_path

    try:
        return resolve_openpyxl_workbook_path(path)
    except BundleOfficeConvertError as e:
        raise EmployeeExcelError(str(e)) from e


@dataclass
class EmployeeRecord:
    """Строка сотрудника из листа rabotnik (ФИО, должность, подразделение, совмещаемая профессия, СНИЛС)."""

    fio: str
    profession: str = ""
    subdivision: str = ""
    profession2: str = ""
    snils: str = ""


def _norm_employee_sort_str(value: str) -> str:
    return (value or "").strip().lower().replace("ё", "е")


def sort_employees_by_fio_alphabet(records: list[EmployeeRecord]) -> None:
    """Сортирует список на месте по ФИО (лексикографически, без учёта регистра; буква ё рядом с е)."""

    records.sort(key=lambda rec: _norm_employee_sort_str(rec.fio))


def sort_employees_by_subdivision_then_fio(records: list[EmployeeRecord]) -> None:
    """Сортирует на месте: по подразделению, затем по ФИО; пустое подразделение — в конце списка."""

    def _key(rec: EmployeeRecord) -> tuple[int, str, str]:
        sub = _norm_employee_sort_str(rec.subdivision)
        fio = _norm_employee_sort_str(rec.fio)
        return (1 if not sub else 0, sub, fio)

    records.sort(key=_key)


@dataclass
class TechVProgramInfo:
    """Строка листа Tech_V: кто утвердил, наименование программы, дата (протокол по техническим вопросам)."""

    approver: str
    program_name: str
    approval_date_raw: str


EMPLOYEES_EXCEL_FILENAME = "Data_base.xlsx"
# Справочник программ (B, V_PROF, PP, SIZ, V). Если файла нет — программы читаются из Data_base.xlsx.
PROGRAMS_EXCEL_FILENAME = "Programs_base.xlsx"


def _is_path_under_office_cache(path: Path) -> bool:
    try:
        from bundle_integration import office_cache_dir

        path.resolve().relative_to(office_cache_dir().resolve())
        return True
    except (ValueError, OSError, ImportError):
        return False


def employees_workbook_writable_path(path: Path) -> Path:
    """
    Куда сохранять правки сотрудников.

    Никогда не пишем в ``.office_cache`` (временный кэш LibreOffice из ODS).
    Не пишем в каталог встроенных ресурсов (bundle/data) — только в каталог данных пользователя.
    Для ``.ods``/``.xls`` — соседний ``.xlsx``; если путь уже из кэша — ``Data_base.xlsx``
    в каталоге данных пользователя.
    """
    path = Path(path)
    if _is_path_under_office_cache(path):
        from app_paths import application_user_dir

        return application_user_dir() / EMPLOYEES_EXCEL_FILENAME
    try:
        from app_paths import application_bundle_dir, application_user_dir

        bundle_root = application_bundle_dir().resolve()
        resolved = path.resolve()
        try:
            resolved.relative_to(bundle_root)
        except ValueError:
            pass
        else:
            return application_user_dir() / EMPLOYEES_EXCEL_FILENAME
    except (OSError, RuntimeError, ImportError):
        pass
    suf = path.suffix.lower()
    if suf in (".xlsx", ".xlsm"):
        return path
    if suf in (".ods", ".xls"):
        return path.with_suffix(".xlsx")
    if suf:
        return path.with_suffix(".xlsx")
    return path.with_name(f"{path.name}.xlsx")


PROGRAM_WORKBOOK_CANONICAL_SHEETS: tuple[str, ...] = (
    "B",
    "V_PROF",
    "PP",
    "SIZ",
    "V",
)
EMPLOYEES_SHEET_NAME = "rabotnik"
EMPLOYEES_ARCHIVE_SHEET_NAME = "rabotnik_archive"
EMPLOYEES_ARCHIVE_SHEET_ALIASES: tuple[str, ...] = (
    "rabotnik_archive",
    "архив",
    "архив сотрудников",
    "archive",
)
EMPLOYEES_SHEET_ALIASES: tuple[str, ...] = (
    "rabotnik",
    "работник",
    "работники",
    "сотрудники",
    "сотрудник",
    "список сотрудников",
    "кадры",
)


@dataclass(frozen=True)
class ArchivedEmployeeEntry:
    """Строка листа архива: номер строки Excel + запись (для восстановления без повторного сопоставления)."""

    row_num: int
    record: EmployeeRecord
    sheet_name: str = EMPLOYEES_ARCHIVE_SHEET_NAME


COMMISSION_SHEET_NAME = "komission"
COMMISSION_SHEET_ALIASES: tuple[str, ...] = (
    "komission",
    "комиссия",
    "комиссия по проверке",
    "commission",
)
# Лист komission: данные с 3-й строки Excel. A — ФИО председателя, B — должность справа от A;
# D — ФИО члена, E — должность справа от D.
COMMISSION_FIRST_DATA_ROW = 3
COMMISSION_COL_CHAIR_FIO = 1  # A
COMMISSION_COL_CHAIR_POSITION = 2  # B
COMMISSION_COL_MEMBER_FIO = 4  # D
COMMISSION_COL_MEMBER_POSITION = 5  # E
COMMISSION_MAX_COL = max(COMMISSION_COL_CHAIR_POSITION, COMMISSION_COL_MEMBER_POSITION)
COMMISSION_MAX_SCAN_ROWS = 500

# Лист Tech_V в файле программ: кто утвердил программу, наименование, дата утверждения (протокол по техническим вопросам).
TECH_V_SHEET_NAME = "Tech_V"
TECH_V_SHEET_ALIASES: tuple[str, ...] = (
    "tech_v",
    "тех_v",
    "TECH_V",
    "Тех_V",
)


def _normalize_excel_header(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", s)


def _header_column_role(header: str) -> str | None:
    if not header:
        return None
    # Одна колонка «Фамилия, Имя, Отчество», «Фамилия, И.» / ФИО.
    if "фамилия" in header and "должн" not in header:
        return "fio"
    if (
        "фио" in header
        or "ф.и.о" in header
        or ("фамилия" in header and "имя" in header)
        or any(x in header for x in ("работник", "сотрудник", "full name", "employee"))
    ):
        return "fio"
    # Раздельные колонки ФИО (если есть в файле).
    if header in ("фамилия",) or (
        header.startswith("фамилия ") and "," not in header and "имя" not in header
    ):
        return "surname"
    if header == "имя" or header.startswith("имя "):
        return "name"
    if "отчество" in header:
        return "patronymic"
    if any(
        x in header
        for x in (
            "профессия",
            "должность",
            "специальность",
            "position",
            "job title",
            "title",
            "должн",
        )
    ) and "п/п" not in header:
        return "profession"
    if any(
        x in header
        for x in (
            "подразделение",
            "наименование подразделения",
            "цех",
            "участок",
            "отдел",
            "department",
            "unit",
            "division",
        )
    ):
        return "subdivision"
    if any(
        x in header
        for x in (
            "снилс",
            "страховой номер",
            "индивидуальный номер",
            "snils",
            "страхов",
            "свидетельств",
        )
    ):
        return "snils"
    return None


def _detect_employee_columns(header_row: tuple[Any, ...]) -> dict[str, int]:
    roles: dict[str, int] = {}
    prof_cols: list[int] = []
    for j, cell in enumerate(header_row):
        hn = _normalize_excel_header(cell)
        if ("совмещ" in hn or "вторая" in hn) and (
            "проф" in hn or "должн" in hn or "спец" in hn
        ):
            if "profession2" not in roles:
                roles["profession2"] = j
            continue
        role = _header_column_role(hn)
        if role == "profession":
            prof_cols.append(j)
        elif role and role not in roles:
            roles[role] = j
    if prof_cols:
        roles["profession"] = prof_cols[0]
        if len(prof_cols) > 1 and "profession2" not in roles:
            roles["profession2"] = prof_cols[1]
    return roles


def _excel_cell_str(row: tuple[Any, ...], index: int) -> str:
    """Строка ячейки; целые float (типичный СНИЛС из Excel) без хвоста ``.0``."""
    if index >= len(row) or row[index] is None:
        return ""
    v = row[index]
    if isinstance(v, bool):
        return str(v).strip()
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).replace("\xa0", " ").strip()


_FIO_HEADER_PLACEHOLDERS = frozenset(
    {
        "фио",
        "фамилия, имя, отчество",
        "фамилия имя отчество",
        "фамилия",
        "имя",
        "отчество",
    }
)


def _is_employee_data_fio(fio: str) -> bool:
    n = _normalize_excel_header(fio)
    return bool(n) and n not in _FIO_HEADER_PLACEHOLDERS


def _find_employee_header_row_index(rows: list[tuple[Any, ...] | None]) -> int:
    """Первая строка с распознанными заголовками сотрудников (до 25 строк листа)."""
    for i, row in enumerate(rows[:25]):
        if not row:
            continue
        cols = _detect_employee_columns(tuple(row))
        if "fio" in cols or ("surname" in cols and "name" in cols):
            return i
    return 0


def _employee_fio_from_row(row: tuple[Any, ...], cols: dict[str, int]) -> str:
    if "fio" in cols:
        return _excel_cell_str(row, cols["fio"])
    parts: list[str] = []
    for key in ("surname", "name", "patronymic"):
        if key in cols:
            part = _excel_cell_str(row, cols[key])
            if part:
                parts.append(part)
    return " ".join(parts)


def _detect_commission_first_data_row(ws: Any) -> int:
    """
    Строка с подписями «ФИО» в колонках A и D → данные со следующей строки.
    Иначе — COMMISSION_FIRST_DATA_ROW (как в старых шаблонах).
    """
    max_r = min(int(ws.max_row or 0), COMMISSION_FIRST_DATA_ROW + 12)
    for r in range(1, max_r + 1):
        row = next(
            ws.iter_rows(
                min_row=r,
                max_row=r,
                min_col=1,
                max_col=COMMISSION_MAX_COL,
                values_only=True,
            ),
            None,
        )
        tup = tuple(row) if row is not None else ()
        a = _normalize_excel_header(_row_value_str(tup, COMMISSION_COL_CHAIR_FIO))
        d = _normalize_excel_header(_row_value_str(tup, COMMISSION_COL_MEMBER_FIO))
        if a == "фио" and d == "фио":
            return r + 1
    return COMMISSION_FIRST_DATA_ROW


def _pick_employee_worksheet(wb: Any, preferred_sheet: str) -> Any:
    """Лист сотрудников: сначала точное имя (без регистра), затем алиасы."""
    try:
        raw_names = wb.sheetnames
    except AttributeError as e:
        raise EmployeeExcelError("Не удалось прочитать список листов книги Excel.") from e
    names_lower: dict[str, str] = {}
    for n in raw_names:
        names_lower[n.lower().strip()] = n
    hints: list[str] = []
    for h in (preferred_sheet,) + EMPLOYEES_SHEET_ALIASES:
        t = h.strip()
        if t and t.lower() not in [x.lower() for x in hints]:
            hints.append(t)
    for h in hints:
        real = names_lower.get(h.lower().strip())
        if real is not None:
            return wb[real]
    avail = ", ".join(raw_names) if raw_names else "(листов нет)"
    raise EmployeeExcelError(
        f"Не найден лист сотрудников. Проверьте имя листа (ожидается «{EMPLOYEES_SHEET_NAME}» или похожее).\n"
        f"Доступные листы: {avail}"
    )


def load_employees_from_excel(path: Path, *, sheet_name: str = EMPLOYEES_SHEET_NAME) -> list[EmployeeRecord]:
    """Читает сотрудников с листа Excel; первая строка — заголовки с ФИО / должностью / подразделением."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise EmployeeExcelError(
            "Не установлен пакет openpyxl. Выполните в папке проекта:\n"
            "  pip install openpyxl"
        ) from e

    if not path.is_file():
        raise EmployeeExcelError(f"Файл не найден:\n{path}")

    path = _workbook_path_for_openpyxl(path)

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise EmployeeExcelError(
            f"Не удалось открыть файл (нужен формат .xlsx или .xlsm, не старый .xls):\n{path}\n\n"
            f"{type(e).__name__}: {e}"
        ) from e
    try:
        ws = _pick_employee_worksheet(wb, sheet_name)
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return []
        header_idx = _find_employee_header_row_index(all_rows)
        header = tuple(all_rows[header_idx])
        cols = _detect_employee_columns(header)
        has_fio_cols = "fio" in cols or ("surname" in cols and "name" in cols)
        if not has_fio_cols:
            col_fio, col_prof, col_sub = 0, 1, 2
            col_prof2 = -1
            col_snils = -1
            use_legacy_cols = True
        else:
            col_prof = cols.get("profession", -1)
            col_sub = cols.get("subdivision", -1)
            col_prof2 = cols.get("profession2", -1)
            col_snils = cols.get("snils", -1)
            use_legacy_cols = False

        out: list[EmployeeRecord] = []
        for row in all_rows[header_idx + 1 :]:
            if not row:
                continue
            tup = tuple(row)
            if use_legacy_cols:
                fio = _excel_cell_str(tup, col_fio)
            else:
                fio = _employee_fio_from_row(tup, cols)
            if not _is_employee_data_fio(fio):
                continue
            p2 = _excel_cell_str(tup, col_prof2) if col_prof2 >= 0 else ""
            sn = _excel_cell_str(tup, col_snils) if col_snils >= 0 else ""
            prof = (
                _excel_cell_str(tup, col_prof)
                if col_prof >= 0
                else (_excel_cell_str(tup, 1) if use_legacy_cols else "")
            )
            sub = (
                _excel_cell_str(tup, col_sub)
                if col_sub >= 0
                else (_excel_cell_str(tup, 2) if use_legacy_cols else "")
            )
            out.append(
                EmployeeRecord(
                    fio=fio,
                    profession=prof,
                    subdivision=sub,
                    profession2=p2,
                    snils=sn,
                )
            )
        return out
    except EmployeeExcelError:
        raise
    except Exception as e:
        _logger.exception("Ошибка чтения листа сотрудников: %s", path)
        raise EmployeeExcelError(
            f"Ошибка при чтении листа сотрудников в файле:\n{path}\n\n{type(e).__name__}: {e}"
        ) from e
    finally:
        wb.close()


def _pick_commission_worksheet(wb: Any) -> Any:
    """Лист комиссии: сначала «komission», затем алиасы (без учёта регистра)."""
    try:
        raw_names = wb.sheetnames
    except AttributeError as e:
        raise EmployeeExcelError("Не удалось прочитать список листов книги Excel.") from e
    names_lower: dict[str, str] = {}
    for n in raw_names:
        names_lower[n.lower().strip()] = n
    hints: list[str] = []
    for h in (COMMISSION_SHEET_NAME,) + COMMISSION_SHEET_ALIASES:
        t = h.strip()
        if t and t.lower() not in [x.lower() for x in hints]:
            hints.append(t)
    for h in hints:
        real = names_lower.get(h.lower().strip())
        if real is not None:
            return wb[real]
    avail = ", ".join(raw_names) if raw_names else "(листов нет)"
    raise EmployeeExcelError(
        f"Не найден лист комиссии. Ожидается «{COMMISSION_SHEET_NAME}» или похожее имя.\n"
        f"Доступные листы: {avail}"
    )


def _row_value_str(row: tuple[Any, ...], col_one_based: int) -> str:
    idx = col_one_based - 1
    if idx < 0 or idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).replace("\r\n", " ").replace("\n", " ").strip()


def load_commission_from_excel(path: Path) -> list[EmployeeRecord]:
    """
    Читает кандидатов в комиссию с листа komission.
    Со строки COMMISSION_FIRST_DATA_ROW: A+должность в B, D+должность в E;
    в списке — уникальные пары ФИО+должность (по строкам: сначала блок A/B, затем D/E).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise EmployeeExcelError(
            "Не установлен пакет openpyxl. Выполните в папке проекта:\n"
            "  pip install openpyxl"
        ) from e

    if not path.is_file():
        raise EmployeeExcelError(f"Файл не найден:\n{path}")

    path = _workbook_path_for_openpyxl(path)

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        _logger.exception("Не удалось открыть Excel (комиссия): %s", path)
        raise EmployeeExcelError(
            f"Не удалось открыть файл (нужен формат .xlsx или .xlsm):\n{path}\n\n"
            f"{type(e).__name__}: {e}"
        ) from e
    try:
        ws = _pick_commission_worksheet(wb)
        first_row = _detect_commission_first_data_row(ws)
        max_r = ws.max_row or 0
        last = min(max_r, first_row + COMMISSION_MAX_SCAN_ROWS - 1)
        if last < first_row:
            return []

        seen_keys: set[str] = set()
        out: list[EmployeeRecord] = []

        def _try_add(fio_col: int, pos_col: int, tup: tuple[Any, ...]) -> None:
            fio = _row_value_str(tup, fio_col)
            if not fio or _normalize_excel_header(fio) == "фио":
                return
            rec = EmployeeRecord(
                fio=fio,
                profession=_row_value_str(tup, pos_col),
            )
            key = employee_unique_key(rec)
            if key in seen_keys:
                return
            seen_keys.add(key)
            out.append(rec)

        for row in ws.iter_rows(
            min_row=first_row,
            max_row=last,
            min_col=1,
            max_col=COMMISSION_MAX_COL,
            values_only=True,
        ):
            tup = tuple(row) if row is not None else ()
            _try_add(COMMISSION_COL_CHAIR_FIO, COMMISSION_COL_CHAIR_POSITION, tup)
            _try_add(COMMISSION_COL_MEMBER_FIO, COMMISSION_COL_MEMBER_POSITION, tup)
        return out
    except EmployeeExcelError:
        raise
    except Exception as e:
        _logger.exception("Ошибка чтения листа комиссии: %s", path)
        raise EmployeeExcelError(
            f"Ошибка при чтении листа комиссии в файле:\n{path}\n\n{type(e).__name__}: {e}"
        ) from e
    finally:
        wb.close()


def _pick_tech_v_worksheet(wb: Any) -> Any:
    try:
        raw_names = wb.sheetnames
    except AttributeError as e:
        raise EmployeeExcelError("Не удалось прочитать список листов книги Excel.") from e
    names_lower: dict[str, str] = {}
    for n in raw_names:
        names_lower[n.lower().strip()] = n
    hints: list[str] = []
    for h in (TECH_V_SHEET_NAME,) + TECH_V_SHEET_ALIASES:
        t = h.strip()
        if t and t.lower() not in [x.lower() for x in hints]:
            hints.append(t)
    for h in hints:
        real = names_lower.get(h.lower().strip())
        if real is not None:
            return wb[real]
    avail = ", ".join(raw_names) if raw_names else "(листов нет)"
    raise EmployeeExcelError(
        f"Не найден лист Tech_V (программы по техническим вопросам). "
        f"Добавьте лист «{TECH_V_SHEET_NAME}» в файл программ.\n"
        f"Доступные листы: {avail}"
    )


def _tech_v_column_map(header_row: tuple[Any, ...]) -> dict[str, int] | None:
    """По строке заголовков: столбцы approver / program / date; None — фиксированные A,B,C."""
    roles: dict[str, int] = {}
    for j, cell in enumerate(header_row):
        hn = _normalize_excel_header(cell)
        if not hn:
            continue
        if "утверд" in hn:
            roles["approver"] = j
        elif "програм" in hn or "наимен" in hn or "тем" in hn or "назван" in hn:
            roles["program"] = j
        elif hn.startswith("дата") or " дата" in f" {hn} ":
            roles["date"] = j
    if "program" in roles and ("approver" in roles or "date" in roles):
        return roles
    if "program" in roles:
        return roles
    return None


def load_all_tech_v_programs_from_excel(path: Path) -> list[TechVProgramInfo]:
    """
    Все строки листа Tech_V с непустым наименованием программы (сверху вниз).
    Заголовки и столбцы — как у load_tech_v_program_from_excel.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise EmployeeExcelError(
            "Не установлен пакет openpyxl. Выполните в папке проекта:\n  pip install openpyxl"
        ) from e

    if not path.is_file():
        raise EmployeeExcelError(f"Файл программ не найден:\n{path}")

    path = _workbook_path_for_openpyxl(path)

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise EmployeeExcelError(
            f"Не удалось открыть файл программ:\n{path}\n\n{type(e).__name__}: {e}"
        ) from e
    try:
        ws = _pick_tech_v_worksheet(wb)
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            raise EmployeeExcelError(f"Лист «{TECH_V_SHEET_NAME}» пуст в файле:\n{path}")

        cmap = _tech_v_column_map(rows[0])
        start = 1 if cmap else 0
        if not cmap:
            cmap = {"approver": 0, "program": 1, "date": 2}

        def cell(row: tuple[Any, ...], role: str) -> str:
            j = cmap.get(role, -1)
            if j < 0:
                return ""
            return _excel_cell_str(row, j)

        out: list[TechVProgramInfo] = []
        for i in range(start, len(rows)):
            tup = rows[i]
            prog = cell(tup, "program").strip()
            if not prog:
                continue
            out.append(
                TechVProgramInfo(
                    approver=cell(tup, "approver").strip(),
                    program_name=prog,
                    approval_date_raw=cell(tup, "date").strip(),
                )
            )

        if not out:
            raise EmployeeExcelError(
                f"На листе «{TECH_V_SHEET_NAME}» нет строки с наименованием программы (столбец программы пуст).\n{path}"
            )
        return out
    except EmployeeExcelError:
        raise
    except Exception as e:
        _logger.exception("Ошибка чтения Tech_V: %s", path)
        raise EmployeeExcelError(
            f"Ошибка при чтении листа Tech_V:\n{path}\n\n{type(e).__name__}: {e}"
        ) from e
    finally:
        wb.close()


def load_tech_v_program_from_excel(path: Path) -> TechVProgramInfo:
    """
    Первая строка с программой на листе Tech_V (для обратной совместимости).
    """
    rows = load_all_tech_v_programs_from_excel(path)
    return rows[0]


def format_fio_filename_surname_initials(fio: str) -> str:
    """
    Краткая подпись для имени файла: «Фамилия И.О.» (например, Иванов И.П.).
    Вход: «Фамилия Имя Отчество», «Фамилия И. О.», одно слово — как есть.
    """
    fio = (fio or "").strip()
    if not fio:
        return ""
    parts = fio.split()
    if len(parts) == 1:
        return parts[0]

    m = re.match(
        r"^(.+?)\s+([А-ЯЁA-Z])\.([А-ЯЁA-Z])\.(\s*)$",
        fio,
        re.IGNORECASE,
    )
    if m:
        return f"{m.group(1).strip()} {m.group(2).upper()}.{m.group(3).upper()}."

    m = re.match(
        r"^(.+?)\s+([А-ЯЁA-Z])\.(\s+)([А-ЯЁA-Z])\.(\s*)$",
        fio,
        re.IGNORECASE,
    )
    if m:
        return f"{m.group(1).strip()} {m.group(2).upper()}.{m.group(4).upper()}."

    m = re.match(r"^(.+?)\s+([А-ЯЁA-Z])\.(\s*)$", fio, re.IGNORECASE)
    if m:
        return f"{m.group(1).strip()} {m.group(2).upper()}."

    fam = parts[0]
    if len(parts) >= 3 and all("." not in p for p in parts):
        name, pat = parts[1], parts[2]
        if name and pat:
            return f"{fam} {name[0].upper()}.{pat[0].upper()}."
    if len(parts) == 2 and "." not in parts[1]:
        name = parts[1]
        if name:
            return f"{fam} {name[0].upper()}."
    return fam


def format_fio_iof(fio: str) -> str:
    """
    ФИО для подписи: «И.О. Фамилия» (инициалы — имя и отчество, затем фамилия).
    Допускает вход «Фамилия Имя Отчество», «Фамилия И.О.», «Фамилия И. О.».
    """
    fio = (fio or "").strip()
    if not fio:
        return ""
    parts = fio.split()
    if len(parts) == 1:
        return parts[0]

    # Фамилия И.О. (без пробела между инициалами)
    m = re.match(
        r"^(.+?)\s+([А-ЯЁA-Z])\.([А-ЯЁA-Z])\.(\s*)$",
        fio,
        re.IGNORECASE,
    )
    if m:
        fam = m.group(1).strip()
        return f"{m.group(2).upper()}.{m.group(3).upper()}. {fam}"

    # Фамилия И. О.
    m = re.match(
        r"^(.+?)\s+([А-ЯЁA-Z])\.(\s+)([А-ЯЁA-Z])\.(\s*)$",
        fio,
        re.IGNORECASE,
    )
    if m:
        fam = m.group(1).strip()
        return f"{m.group(2).upper()}.{m.group(4).upper()}. {fam}"

    # Фамилия И.
    m = re.match(r"^(.+?)\s+([А-ЯЁA-Z])\.(\s*)$", fio, re.IGNORECASE)
    if m:
        fam = m.group(1).strip()
        return f"{m.group(2).upper()}. {fam}"

    # Фамилия Имя Отчество (без точек в частях)
    if len(parts) == 3 and all("." not in p for p in parts):
        fam, name, pat = parts
        if name and pat:
            return f"{name[0].upper()}.{pat[0].upper()}. {fam}"

    if len(parts) == 2 and "." not in parts[1]:
        fam, name = parts
        if name:
            return f"{name[0].upper()}. {fam}"

    return fio


def format_person_iof_line(fio: str, profession: str = "") -> str:
    """Строка для блока подписей: И.О. Фамилия, должность (именительный падеж как в базе)."""
    s = format_fio_iof(fio)
    p = (profession or "").strip()
    if s and p:
        return f"{s}, {p}"
    return s or p


def listbox_label_for_employee(
    rec: EmployeeRecord, *, grouped_by_subdivision: bool = False
) -> str:
    extra = ""
    if rec.profession2:
        extra = f" + {rec.profession2}"
    if grouped_by_subdivision:
        prof = (rec.profession or "").strip() or "—"
        return f"  {rec.fio} — {prof}{extra}"
    if rec.profession:
        return f"{rec.fio} — {rec.profession}{extra}"
    return rec.fio


def subdivision_group_key(subdivision: str) -> str:
    """Ключ группы подразделения для сворачивания списка в интерфейсе."""
    s = (subdivision or "").strip().lower().replace("ё", "е")
    return s if s else "__no_sub__"


def listbox_subdivision_header(
    subdivision: str,
    employee_count: int,
    *,
    collapsed: bool = False,
) -> str:
    """Заголовок группы в списке сотрудников (клик — свернуть/развернуть)."""
    sub = (subdivision or "").strip() or "(без подразделения)"
    n = max(0, int(employee_count))
    mark = "▸" if collapsed else "▾"
    hint = " — свернуто, щёлкните чтобы развернуть" if collapsed else ""
    return f"{mark} {sub}  ({n}){hint}"


def employee_unique_key(rec: EmployeeRecord) -> str:
    return "|".join(
        (
            _norm_employee_match_part(rec.fio),
            _norm_employee_match_part(rec.profession),
            _norm_employee_match_part(rec.subdivision),
            _norm_employee_match_part(rec.profession2),
        )
    )


def employee_archive_restore_key(rec: EmployeeRecord) -> str:
    """Ключ восстановления из архива: ФИО + должность (как в списке «Архив…»)."""
    return "|".join(
        (
            _norm_employee_match_part(rec.fio),
            _norm_employee_match_part(rec.profession),
        )
    )


def employee_archive_match_key(rec: EmployeeRecord) -> str:
    """Ключ для поиска в архиве с учётом подразделения (архивирование и пр.)."""
    return "|".join(
        (
            _norm_employee_match_part(rec.fio),
            _norm_employee_match_part(rec.profession),
            _norm_employee_match_part(rec.subdivision),
        )
    )


def _norm_employee_match_part(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().lower().replace("ё", "е"))


def write_template_data_base_workbook(path: Path) -> None:
    """
    Пустая книга Data_base.xlsx: лист сотрудников (заголовки) и лист комиссии (пояснение + данные с 3-й строки).
    """
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise EmployeeExcelError(
            "Не установлен пакет openpyxl. Выполните: pip install openpyxl"
        ) from e

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = EMPLOYEES_SHEET_NAME
    ws.append(
        [
            "№ п/п",
            "Таб.№",
            "Фамилия, Имя, Отчество",
            "Подразделение",
            "Должность",
            "№ страхового свидетельства",
        ],
    )
    ws2 = wb.create_sheet(COMMISSION_SHEET_NAME)
    ws2["A1"] = "Председатель"
    ws2["D1"] = "Члены  комиссии"
    ws2.append(["ФИО", "Должность", None, "ФИО", "Должность"])
    wb.save(path)


def _workbook_sheet_lookup(wb: Any) -> dict[str, str]:
    return {n.lower(): n for n in wb.sheetnames}


def _source_sheet_name_for_program(wb: Any, canonical: str) -> str | None:
    m = _workbook_sheet_lookup(wb)
    if canonical.lower() in m:
        return m[canonical.lower()]
    if canonical == "V":
        for alias in ("v", "в"):
            if alias in m:
                return m[alias]
    return None


def copy_program_sheets_from_workbook(source: Path, dest: Path) -> list[str]:
    """
    Копирует листы программ из объединённого Data_base (или другого файла) в отдельную книгу Programs_base.
    Возвращает список скопированных канонических имён листов.
    """
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as e:
        raise EmployeeExcelError(
            "Не установлен пакет openpyxl. Выполните: pip install openpyxl"
        ) from e

    source = Path(source)
    dest = Path(dest)
    if not source.is_file():
        raise EmployeeExcelError(f"Файл не найден:\n{source}")

    source = _workbook_path_for_openpyxl(source)

    swb = load_workbook(source, data_only=True, read_only=True)
    try:
        dwb = Workbook()
        if dwb.active is not None:
            dwb.remove(dwb.active)
        copied: list[str] = []
        for canon in PROGRAM_WORKBOOK_CANONICAL_SHEETS:
            src_name = _source_sheet_name_for_program(swb, canon)
            if not src_name:
                continue
            ws_src = swb[src_name]
            ws_dst = dwb.create_sheet(canon)
            for row in ws_src.iter_rows(values_only=True):
                ws_dst.append(list(row) if row is not None else [])
            copied.append(canon)
    finally:
        swb.close()

    if not copied:
        raise EmployeeExcelError(
            "В выбранном файле нет листов программ (ожидаются: B, V_PROF, PP, SIZ, V)."
        )

    dest.parent.mkdir(parents=True, exist_ok=True)
    dwb.save(dest)
    return copied


def split_combined_employees_workbook(
    employees_path: Path,
    programs_path: Path | None = None,
    *,
    backup_employees: bool = True,
) -> tuple[list[str], Path]:
    """
    Вынести листы программ из объединённого Data_base в Programs_base.xlsx и удалить их из книги сотрудников.

    Сначала (при backup_employees) сохраняется копия «имя_before_split.xlsx» с полным содержимым до разбиения.
    Значения на листах программ копируются как в copy_program_sheets_from_workbook (без формул).
    """
    employees_path = Path(employees_path).expanduser().resolve()
    programs_path = (
        Path(programs_path).expanduser().resolve()
        if programs_path is not None
        else employees_path.parent / PROGRAMS_EXCEL_FILENAME
    )
    if not employees_path.is_file():
        raise EmployeeExcelError(f"Файл не найден:\n{employees_path}")

    if backup_employees:
        bak = employees_path.with_name(f"{employees_path.stem}_before_split{employees_path.suffix}")
        shutil.copy2(employees_path, bak)

    copied = copy_program_sheets_from_workbook(employees_path, programs_path)

    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise EmployeeExcelError(
            "Не установлен пакет openpyxl. Выполните: pip install openpyxl"
        ) from e

    wb = load_workbook(employees_path, data_only=False)
    to_remove: list[str] = []
    for canon in PROGRAM_WORKBOOK_CANONICAL_SHEETS:
        sn = _source_sheet_name_for_program(wb, canon)
        if sn and sn not in to_remove:
            to_remove.append(sn)
    for sn in to_remove:
        wb.remove(wb[sn])
    if not wb.sheetnames:
        raise EmployeeExcelError(
            "После удаления листов программ в книге не осталось листов — файл не сохранён (восстановите из копии _before_split)."
        )
    wb.save(employees_path)
    return copied, programs_path


def write_template_programs_workbook(path: Path) -> None:
    """Пустая книга Programs_base.xlsx: листы B, V_PROF, PP, SIZ, V с поясняющими заголовками."""
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise EmployeeExcelError(
            "Не установлен пакет openpyxl. Выполните: pip install openpyxl"
        ) from e

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "B"
    ws.append(
        [
            "(опц.) примечание",
            "Наименование программы «Б» (столбец 2)",
            "Объём, ч (столбец 3, опц.)",
        ]
    )
    vp = wb.create_sheet("V_PROF")
    hdr = ["Должность (A)", "Якорь «Б» (2)", "Якорь ПП (3)", "Якорь СИЗ (4)"]
    hdr.extend([f"Фрагм. «В» ({i})" for i in range(5, 23)])
    vp.append(hdr)
    pp = wb.create_sheet("PP")
    pp.append(["…", "Наименование ПП для таблицы (столбец 2)", "Объём, ч (ст. 3, опц.)"])
    siz = wb.create_sheet("SIZ")
    siz.append(["…", "Наименование СИЗ (столбец 2)", "Объём, ч (ст. 3, опц.)"])
    v = wb.create_sheet("V")
    v.append(
        [
            "ID в гос. реестре (A)",
            "Сопоставление с V_PROF (B)",
            "Наименование после проверки (C)",
            "Объём, ч (столбец D, опц.)",
        ]
    )
    wb.save(path)


@dataclass
class _EmployeeSheetColumns:
    """Разметка листа сотрудников для записи в Excel (номера строк/столбцов — как в openpyxl)."""

    header_row: int
    cols: dict[str, int]
    use_legacy: bool
    col_fio: int
    col_prof: int
    col_sub: int
    col_prof2: int
    col_snils: int
    serial_col: int
    max_col: int


def _detect_serial_column(header_row: tuple[Any, ...]) -> int:
    """Только колонка «№ п/п», не Таб.№ и не СНИЛС."""
    for j, cell in enumerate(header_row):
        hn = _normalize_excel_header(cell)
        if "п/п" in hn:
            return j
    return -1


def _analyze_employee_worksheet(ws: Any) -> _EmployeeSheetColumns:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        header_idx = 0
        header: tuple[Any, ...] = ()
        cols: dict[str, int] = {}
        use_legacy = True
    else:
        header_idx = _find_employee_header_row_index(rows)
        header = tuple(rows[header_idx])
        cols = _detect_employee_columns(header)
        use_legacy = not ("fio" in cols or ("surname" in cols and "name" in cols))
    header_row = header_idx + 1
    serial_col = _detect_serial_column(header) if header else -1
    max_col = max(len(header), 6) if header else 6
    return _EmployeeSheetColumns(
        header_row=header_row,
        cols=cols,
        use_legacy=use_legacy,
        col_fio=0,
        col_prof=cols.get("profession", 1) if not use_legacy else 1,
        col_sub=cols.get("subdivision", 2) if not use_legacy else 2,
        col_prof2=cols.get("profession2", -1) if not use_legacy else -1,
        col_snils=cols.get("snils", -1) if not use_legacy else -1,
        serial_col=serial_col,
        max_col=max_col,
    )


def _employee_record_from_row_values(
    tup: tuple[Any, ...],
    layout: _EmployeeSheetColumns,
) -> EmployeeRecord | None:
    if layout.use_legacy:
        fio = _excel_cell_str(tup, layout.col_fio)
    else:
        fio = _employee_fio_from_row(tup, layout.cols)
    if not _is_employee_data_fio(fio):
        return None
    p2 = _excel_cell_str(tup, layout.col_prof2) if layout.col_prof2 >= 0 else ""
    sn = _excel_cell_str(tup, layout.col_snils) if layout.col_snils >= 0 else ""
    prof = _excel_cell_str(tup, layout.col_prof) if layout.col_prof >= 0 else ""
    sub = _excel_cell_str(tup, layout.col_sub) if layout.col_sub >= 0 else ""
    return EmployeeRecord(fio=fio, profession=prof, subdivision=sub, profession2=p2, snils=sn)


def _employee_records_match(a: EmployeeRecord, b: EmployeeRecord) -> bool:
    return employee_unique_key(a) == employee_unique_key(b)


def _employee_archive_records_match(a: EmployeeRecord, b: EmployeeRecord) -> bool:
    return employee_archive_match_key(a) == employee_archive_match_key(b)


def _employee_archive_restore_match(a: EmployeeRecord, b: EmployeeRecord) -> bool:
    return employee_archive_restore_key(a) == employee_archive_restore_key(b)


def _collect_employee_rows_from_sheet(
    ws: Any, layout: _EmployeeSheetColumns
) -> list[tuple[int, EmployeeRecord]]:
    """Строки сотрудников: (номер строки Excel, запись). Не опирается на ws.max_row."""
    out: list[tuple[int, EmployeeRecord]] = []
    for row_num, row_vals in enumerate(
        ws.iter_rows(min_row=layout.header_row + 1, values_only=True),
        start=layout.header_row + 1,
    ):
        if not row_vals:
            continue
        if not any(v is not None and str(v).strip() for v in row_vals):
            continue
        rec = _employee_record_from_row_values(tuple(row_vals), layout)
        if rec is not None:
            out.append((row_num, rec))
    return out


def _last_employee_data_row(ws: Any, layout: _EmployeeSheetColumns) -> int:
    """Последняя строка с данными сотрудника (не ws.max_row)."""
    rows = _collect_employee_rows_from_sheet(ws, layout)
    if not rows:
        return layout.header_row
    return max(row_num for row_num, _rec in rows)


def _append_employee_record(
    ws: Any,
    layout: _EmployeeSheetColumns,
    record: EmployeeRecord,
    *,
    assign_serial: bool = False,
) -> int:
    """Добавить запись в конец таблицы; № п/п не заполняется (assign_serial=False)."""
    new_row = _last_employee_data_row(ws, layout) + 1
    if assign_serial and layout.serial_col >= 0:
        serial = _next_serial_number(ws, layout)
        if serial is not None:
            ws.cell(row=new_row, column=layout.serial_col + 1, value=serial)
    _write_employee_cells(ws, new_row, record, layout)
    return new_row


def _clear_serial_cell(ws: Any, row_num: int, layout: _EmployeeSheetColumns) -> None:
    """Не заполняем и не сохраняем № п/п — только данные сотрудника."""
    if layout.serial_col >= 0:
        ws.cell(row=row_num, column=layout.serial_col + 1, value="")


def _copy_employee_row_clear_serial(
    ws_src: Any,
    src_row: int,
    ws_dst: Any,
    dst_row: int,
    *,
    src_max_col: int,
    dst_layout: _EmployeeSheetColumns,
) -> None:
    """Копия строки Excel с сохранением Таб.№ и прочих колонок; № п/п очищается."""
    _copy_excel_row(ws_src, src_row, ws_dst, dst_row, src_max_col)
    _clear_serial_cell(ws_dst, dst_row, dst_layout)


def _write_employee_cells(
    ws: Any,
    row_num: int,
    record: EmployeeRecord,
    layout: _EmployeeSheetColumns,
) -> None:
    if layout.use_legacy:
        ws.cell(row=row_num, column=layout.col_fio + 1, value=(record.fio or "").strip())
        ws.cell(row=row_num, column=layout.col_prof + 1, value=(record.profession or "").strip())
        ws.cell(row=row_num, column=layout.col_sub + 1, value=(record.subdivision or "").strip())
        if layout.col_prof2 >= 0:
            ws.cell(row=row_num, column=layout.col_prof2 + 1, value=(record.profession2 or "").strip())
        if layout.col_snils >= 0:
            ws.cell(row=row_num, column=layout.col_snils + 1, value=(record.snils or "").strip())
        _clear_serial_cell(ws, row_num, layout)
        return
    if "fio" in layout.cols:
        ws.cell(row=row_num, column=layout.cols["fio"] + 1, value=(record.fio or "").strip())
    elif "surname" in layout.cols and "name" in layout.cols:
        parts = (record.fio or "").split()
        ws.cell(row=row_num, column=layout.cols["surname"] + 1, value=parts[0] if parts else "")
        ws.cell(
            row=row_num,
            column=layout.cols["name"] + 1,
            value=parts[1] if len(parts) > 1 else "",
        )
        if "patronymic" in layout.cols:
            ws.cell(
                row=row_num,
                column=layout.cols["patronymic"] + 1,
                value=parts[2] if len(parts) > 2 else "",
            )
    if layout.col_prof >= 0:
        ws.cell(row=row_num, column=layout.col_prof + 1, value=(record.profession or "").strip())
    if layout.col_sub >= 0:
        ws.cell(row=row_num, column=layout.col_sub + 1, value=(record.subdivision or "").strip())
    if layout.col_prof2 >= 0:
        ws.cell(row=row_num, column=layout.col_prof2 + 1, value=(record.profession2 or "").strip())
    if layout.col_snils >= 0:
        ws.cell(row=row_num, column=layout.col_snils + 1, value=(record.snils or "").strip())
    _clear_serial_cell(ws, row_num, layout)


def _next_serial_number(ws: Any, layout: _EmployeeSheetColumns) -> int | None:
    if layout.serial_col < 0:
        return None
    best = 0
    for row in ws.iter_rows(
        min_row=layout.header_row + 1,
        max_row=ws.max_row or layout.header_row,
        values_only=True,
    ):
        if not row:
            continue
        raw = row[layout.serial_col] if layout.serial_col < len(row) else None
        try:
            n = int(str(raw).strip())
        except (TypeError, ValueError, AttributeError):
            continue
        best = max(best, n)
    return best + 1


def _copy_excel_row(ws_src: Any, src_row: int, ws_dst: Any, dst_row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        ws_dst.cell(row=dst_row, column=c, value=ws_src.cell(row=src_row, column=c).value)


def _backup_workbook_before_edit(path: Path, *, keep: int = 5) -> Path | None:
    """Копия перед правкой; хранит до ``keep`` файлов ``*_before_edit`` / ``*_before_edit.N``."""
    path = Path(path)
    if not path.is_file():
        return None
    keep = max(1, int(keep))
    primary = path.with_name(f"{path.stem}_before_edit{path.suffix}")
    slots = [primary] + [
        path.with_name(f"{path.stem}_before_edit.{i}{path.suffix}")
        for i in range(1, keep)
    ]
    # Сдвиг в сторону «старее»: slots[i-1] → slots[i]
    for i in range(len(slots) - 1, 0, -1):
        prev, cur = slots[i - 1], slots[i]
        if not prev.is_file():
            continue
        try:
            if cur.is_file():
                cur.unlink()
        except OSError:
            pass
        try:
            prev.replace(cur)
        except OSError:
            try:
                shutil.copy2(prev, cur)
                prev.unlink(missing_ok=True)
            except OSError:
                pass
    shutil.copy2(path, primary)
    return primary


def _open_employees_workbook_for_edit(path: Path) -> Any:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise EmployeeExcelError(
            "Не установлен пакет openpyxl. Выполните: pip install openpyxl"
        ) from e
    path = Path(path)
    if not path.is_file():
        write_template_data_base_workbook(path)
    open_path = _workbook_path_for_openpyxl(path)
    try:
        # data_only=False: при сохранении не теряются формулы на других листах книги.
        return load_workbook(open_path, read_only=False, data_only=False)
    except PermissionError as e:
        raise EmployeeExcelError(
            f"Файл занят другой программой (закройте Excel и повторите):\n{path}"
        ) from e
    except Exception as e:
        raise EmployeeExcelError(
            f"Не удалось открыть файл для записи:\n{path}\n\n{type(e).__name__}: {e}"
        ) from e


def _find_archive_worksheet(
    wb: Any,
    *,
    preferred_sheet: str = EMPLOYEES_ARCHIVE_SHEET_NAME,
) -> tuple[Any, str] | tuple[None, None]:
    """Лист архива сотрудников, если есть (без создания нового)."""
    names_lower = {n.lower().strip(): n for n in wb.sheetnames}
    for hint in (preferred_sheet,) + EMPLOYEES_ARCHIVE_SHEET_ALIASES:
        real = names_lower.get(hint.lower().strip())
        if real is not None:
            return wb[real], real
    return None, None


def _pick_archive_worksheet(wb: Any, main_ws: Any, layout: _EmployeeSheetColumns) -> Any:
    found, _name = _find_archive_worksheet(wb)
    if found is not None:
        return found
    ws = wb.create_sheet(EMPLOYEES_ARCHIVE_SHEET_NAME)
    for c in range(1, layout.max_col + 1):
        ws.cell(row=1, column=c, value=main_ws.cell(row=layout.header_row, column=c).value)
    return ws


def _save_employees_workbook(wb: Any, path: Path) -> None:
    path = Path(path)
    save_path = employees_workbook_writable_path(path)
    save_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(save_path)
    except PermissionError as e:
        raise EmployeeExcelError(
            f"Не удалось сохранить файл (закройте Excel и повторите):\n{save_path}"
        ) from e
    except OSError as e:
        raise EmployeeExcelError(f"Не удалось сохранить файл:\n{save_path}\n\n{e}") from e


def load_archived_employee_entries_from_excel(
    path: Path,
    *,
    sheet_name: str = EMPLOYEES_ARCHIVE_SHEET_NAME,
) -> list[ArchivedEmployeeEntry]:
    """Сотрудники с листа архива с номерами строк Excel (для восстановления)."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise EmployeeExcelError(
            "Не установлен пакет openpyxl. Выполните: pip install openpyxl"
        ) from e
    path = Path(path)
    if not path.is_file():
        return []
    open_path = _workbook_path_for_openpyxl(path)
    wb = load_workbook(open_path, read_only=True, data_only=True)
    try:
        ws, real_name = _find_archive_worksheet(wb, preferred_sheet=sheet_name)
        if ws is None or real_name is None:
            return []
        layout = _analyze_employee_worksheet(ws)
        return [
            ArchivedEmployeeEntry(row_num=row_num, record=rec, sheet_name=real_name)
            for row_num, rec in _collect_employee_rows_from_sheet(ws, layout)
        ]
    finally:
        wb.close()


def load_archived_employees_from_excel(
    path: Path,
    *,
    sheet_name: str = EMPLOYEES_ARCHIVE_SHEET_NAME,
) -> list[EmployeeRecord]:
    """Сотрудники с листа архива (rabotnik_archive)."""
    return [e.record for e in load_archived_employee_entries_from_excel(path, sheet_name=sheet_name)]


def employee_rows_for_excel_add(record: EmployeeRecord) -> list[EmployeeRecord]:
    """
    Строки для записи в Data_base: при совмещаемой профессии — две строки
    (основная должность и совмещение), как при формировании блока «Б» протокола.
    """
    fio = (record.fio or "").strip()
    sub = (record.subdivision or "").strip()
    sn = (record.snils or "").strip()
    main = (record.profession or "").strip()
    extra = (record.profession2 or "").strip()
    if not extra or _norm_employee_sort_str(extra) == _norm_employee_sort_str(main):
        return [
            EmployeeRecord(
                fio=fio,
                profession=main,
                subdivision=sub,
                profession2="",
                snils=sn,
            )
        ]
    return [
        EmployeeRecord(
            fio=fio,
            profession=main,
            subdivision=sub,
            profession2="",
            snils=sn,
        ),
        EmployeeRecord(
            fio=fio,
            profession=extra,
            subdivision=sub,
            profession2="",
            snils=sn,
        ),
    ]


def add_employee_to_excel(
    path: Path,
    record: EmployeeRecord,
    *,
    backup: bool = True,
) -> int:
    """Добавить строки сотрудника на лист rabotnik; при profession2 — две строки. Возвращает число добавленных строк."""
    fio = (record.fio or "").strip()
    if not fio:
        raise EmployeeExcelError("ФИО сотрудника не может быть пустым.")
    rows_to_add = employee_rows_for_excel_add(record)
    path = Path(path)
    if backup and path.is_file():
        _backup_workbook_before_edit(path)
    wb = _open_employees_workbook_for_edit(path)
    try:
        ws = _pick_employee_worksheet(wb, EMPLOYEES_SHEET_NAME)
        layout = _analyze_employee_worksheet(ws)
        if layout.header_row == 1 and (ws.max_row or 0) <= 1 and not layout.cols:
            wb.close()
            write_template_data_base_workbook(path)
            wb = _open_employees_workbook_for_edit(path)
            ws = _pick_employee_worksheet(wb, EMPLOYEES_SHEET_NAME)
            layout = _analyze_employee_worksheet(ws)
        for rec in rows_to_add:
            _append_employee_record(ws, layout, rec, assign_serial=False)
        _save_employees_workbook(wb, path)
    finally:
        wb.close()
    return len(rows_to_add)


def archive_employees_in_excel(
    path: Path,
    records: list[EmployeeRecord],
    *,
    backup: bool = True,
) -> int:
    """Перенести сотрудников с листа rabotnik на лист rabotnik_archive."""
    if not records:
        return 0
    path = Path(path)
    if backup and path.is_file():
        _backup_workbook_before_edit(path)
    wb = _open_employees_workbook_for_edit(path)
    moved = 0
    try:
        ws = _pick_employee_worksheet(wb, EMPLOYEES_SHEET_NAME)
        layout = _analyze_employee_worksheet(ws)
        archive_ws = _pick_archive_worksheet(wb, ws, layout)
        targets = list(records)
        main_rows = _collect_employee_rows_from_sheet(ws, layout)
        for row_num, rec in reversed(main_rows):
            if not targets:
                break
            match_idx = next(
                (i for i, t in enumerate(targets) if _employee_archive_records_match(t, rec)),
                None,
            )
            if match_idx is None:
                continue
            targets.pop(match_idx)
            archive_layout = _analyze_employee_worksheet(archive_ws)
            dst_row = _last_employee_data_row(archive_ws, archive_layout) + 1
            _copy_employee_row_clear_serial(
                ws,
                row_num,
                archive_ws,
                dst_row,
                src_max_col=layout.max_col,
                dst_layout=archive_layout,
            )
            ws.delete_rows(row_num, 1)
            moved += 1
        if moved == 0:
            raise EmployeeExcelError(
                "Не найдены выбранные сотрудники в файле Excel (возможно, список устарел — "
                "нажмите «Обновить базы с диска»)."
            )
        # Частичный перенос допустим: вызывающий код сравнивает returned с len(records).
        _save_employees_workbook(wb, path)
    finally:
        wb.close()
    return moved


def restore_archived_employee_entries(
    path: Path,
    entries: list[ArchivedEmployeeEntry],
    *,
    backup: bool = True,
) -> int:
    """Вернуть на rabotnik строки архива по номерам строк Excel (как в диалоге «Архив…»)."""
    if not entries:
        return 0
    path = Path(path)
    if backup and path.is_file():
        _backup_workbook_before_edit(path)
    wb = _open_employees_workbook_for_edit(path)
    restored = 0
    try:
        ws = _pick_employee_worksheet(wb, EMPLOYEES_SHEET_NAME)
        main_layout = _analyze_employee_worksheet(ws)
        by_sheet: dict[str, list[ArchivedEmployeeEntry]] = {}
        for entry in entries:
            by_sheet.setdefault(entry.sheet_name, []).append(entry)
        for sheet_name, sheet_entries in by_sheet.items():
            archive_ws, real_name = _find_archive_worksheet(wb, preferred_sheet=sheet_name)
            if archive_ws is None or real_name is None:
                continue
            archive_layout = _analyze_employee_worksheet(archive_ws)
            for entry in sorted(sheet_entries, key=lambda e: e.row_num, reverse=True):
                row_num = entry.row_num
                if row_num <= archive_layout.header_row:
                    continue
                row_vals = next(
                    archive_ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True),
                    None,
                )
                if not row_vals:
                    continue
                if _employee_record_from_row_values(tuple(row_vals), archive_layout) is None:
                    continue
                dst_row = _last_employee_data_row(ws, main_layout) + 1
                _copy_employee_row_clear_serial(
                    archive_ws,
                    row_num,
                    ws,
                    dst_row,
                    src_max_col=archive_layout.max_col,
                    dst_layout=main_layout,
                )
                archive_ws.delete_rows(row_num, 1)
                restored += 1
        if restored == 0:
            raise EmployeeExcelError("Не найдены выбранные записи в листе архива.")
        _save_employees_workbook(wb, path)
    finally:
        wb.close()
    return restored


def restore_employees_from_archive(
    path: Path,
    records: list[EmployeeRecord],
    *,
    backup: bool = True,
) -> int:
    """Вернуть сотрудников с листа архива на rabotnik (сопоставление: ФИО + должность)."""
    if not records:
        return 0
    path = Path(path)
    if backup and path.is_file():
        _backup_workbook_before_edit(path)
    wb = _open_employees_workbook_for_edit(path)
    restored = 0
    try:
        ws = _pick_employee_worksheet(wb, EMPLOYEES_SHEET_NAME)
        main_layout = _analyze_employee_worksheet(ws)
        archive_ws, _archive_name = _find_archive_worksheet(wb)
        if archive_ws is None:
            raise EmployeeExcelError("Лист архива не найден (rabotnik_archive).")
        archive_layout = _analyze_employee_worksheet(archive_ws)
        targets = list(records)
        archive_rows = _collect_employee_rows_from_sheet(archive_ws, archive_layout)
        to_restore: list[tuple[int, EmployeeRecord]] = []
        for row_num, rec in archive_rows:
            if not targets:
                break
            match_idx = next(
                (i for i, t in enumerate(targets) if _employee_archive_restore_match(t, rec)),
                None,
            )
            if match_idx is None:
                continue
            targets.pop(match_idx)
            to_restore.append((row_num, rec))
        if not to_restore:
            raise EmployeeExcelError(
                "Не найдены выбранные записи в листе архива (ФИО и должность)."
            )
        if targets:
            raise EmployeeExcelError(
                f"Не все выбранные записи найдены в архиве "
                f"(не найдено: {len(targets)} из {len(records)})."
            )
        for row_num, _rec in sorted(to_restore, key=lambda item: item[0], reverse=True):
            dst_row = _last_employee_data_row(ws, main_layout) + 1
            _copy_employee_row_clear_serial(
                archive_ws,
                row_num,
                ws,
                dst_row,
                src_max_col=archive_layout.max_col,
                dst_layout=main_layout,
            )
            archive_ws.delete_rows(row_num, 1)
            restored += 1
        _save_employees_workbook(wb, path)
    finally:
        wb.close()
    return restored

