# -*- coding: utf-8 -*-
"""
Выгрузка для реестра Минтруда / АКОТ на основе официального Excel-шаблона XSD
(файл «Шаблон_Минтруд_XSD_УМН.xlsx», лист «Шаблон»). Сопоставление XML/XSD сохраняется: каркас книги из шаблона, данные листа — из openpyxl.
Рядом с Excel
создаётся XML RegistrySet для загрузки на сайт (то же имя, расширение .xml).

Для записей с метаданными: «Б» — лист B; «ПП»/«СИЗ» — листы PP/SIZ (одно имя на протокол);
ID — с листа V. «В» — матрица V_PROF («Да») + текст из столбца B листа V; ID — по совпадению с B.
Должность — из persons_raw журнала или Excel по ФИО; для программы «В» — должность из протокола.
Старые записи без meta — одна строка с полем topic.
СНИЛС в шаблоне — из файла сотрудников по ФИО (в meta может быть снимок для справки).
«Тест пройден»: 1 / 0.
"""

from __future__ import annotations

import json
import logging
import re
import shutil
import zipfile
import xml.etree.ElementTree as ET
from collections.abc import Callable
from datetime import date, datetime
from pathlib import Path
from typing import Any

from app_paths import application_bundle_dir
from program_keys import ProgramKey, parse_program_key
from employees_io import EmployeeRecord, EmployeeExcelError, load_employees_from_excel
from excel_data_cache import (
    get_cached_b_program_title,
    get_cached_pp_table_title,
    get_cached_siz_table_title,
    get_cached_v_registry_rows,
)
from v_program_registry_match import match_v_registry_fragment

_log = logging.getLogger(__name__)


def _registry_title_and_id_from_v(
    text: str, v_rows: list[tuple[str, str, str, float | None]]
) -> tuple[str, str]:
    """Совпадение строки с колонкой B листа V → (наименование из C/B, ID из A); иначе исходный текст и пустой ID."""
    s = (text or "").strip()
    if not s or not v_rows:
        return s, ""
    m = match_v_registry_fragment(s, v_rows)
    if m:
        return (m[0] or "").strip(), (m[1] or "").strip()
    return s, ""


def _registry_id_from_v(
    text: str, v_rows: list[tuple[str, str, str, float | None]]
) -> str:
    """ID программы (столбец A листа V) при совпадении наименования с листа B/PP/SIZ."""
    s = (text or "").strip()
    if not s or not v_rows:
        return ""
    m = match_v_registry_fragment(s, v_rows)
    return (m[1] or "").strip() if m else ""


# Актуальное имя шаблона и запасные варианты.
MINTRUD_TEMPLATE_FILENAMES: tuple[str, ...] = (
    "Шаблон_Минтруд_XSD_УМН.xlsx",
    "!! Шаблон_Минтруд_XSD_УМН _ общ+.xlsx",
    "Шаблон_Минтруд_XSD_УМН _ общ+.xlsx",
)
MINTRUD_TEMPLATE_SHEET = "Шаблон"
MINTRUD_XSD_FILENAME = "educated_person_import_v1.0.9.xsd"
HEADER_ROW = 1

_DATE_DMY_RE = re.compile(r"^(\d{1,2})[./](\d{1,2})[./](\d{4})$")
_DATE_ISO_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")


def format_mintrud_iso_date(value: object) -> str:
    """Дата для XSD/сайта Минтруда: YYYY-MM-DD (например 2026-05-20)."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return ""
    if _DATE_ISO_RE.match(s):
        try:
            return date.fromisoformat(s).isoformat()
        except ValueError:
            return ""
    m = _DATE_DMY_RE.match(s)
    if m:
        try:
            return date(
                int(m.group(3)), int(m.group(2)), int(m.group(1))
            ).isoformat()
        except ValueError:
            return ""
    return ""


def parse_mintrud_iso_date(value: object) -> date | None:
    iso = format_mintrud_iso_date(value)
    if not iso:
        return None
    try:
        return date.fromisoformat(iso)
    except ValueError:
        return None


# Порядок полей во внутренней строке (совпадает с логикой листа «Шаблон» Excel).
MINTRUD_XML_FIELD_ORDER: tuple[str, ...] = (
    "program_name",
    "registry_program_id",
    "last_name",
    "first_name",
    "patronymic",
    "snils",
    "position",
    "doc_date",
    "protocol_no",
    "test_passed",
    "inn_employer",
    "employer_name",
    "inn_org2",
    "org2_name",
)

FIELD_BY_HEADER_NORM: dict[str, str] = {
    "наименование программы в гос. реестре": "program_name",
    "id программы в гос. реестре": "registry_program_id",
    "фамилия": "last_name",
    "имя": "first_name",
    "отчество": "patronymic",
    "снилс": "snils",
    "должность": "position",
    "дата в удостоверении": "doc_date",
    "номер протокола": "protocol_no",
    "тест пройден": "test_passed",
    "инн работодателя": "inn_employer",
    "название работодателя": "employer_name",
    "инн организации2": "inn_org2",
    "наименование организации2": "org2_name",
}


def mintrud_template_path() -> Path | None:
    base = application_bundle_dir()
    for name in MINTRUD_TEMPLATE_FILENAMES:
        p = base / name
        if p.is_file():
            return p
    return None


def extract_protocol_number_from_plain_text(text: str) -> str:
    if not text:
        return ""
    for raw in text.splitlines():
        line = raw.replace("\xa0", " ").strip()
        if not line:
            continue
        m = re.search(
            r"(?i)протокол\s*[№#]?\s*(.+)",
            line,
        )
        if not m:
            continue
        tail = m.group(1).strip()
        tail = re.split(r"\s{2,}", tail)[0]
        return tail[:200]
    return ""


def split_journal_fio_field(fio: str) -> list[str]:
    s = (fio or "").strip()
    if not s:
        return [""]
    parts = [p.strip() for p in re.split(r",\s*", s) if p.strip()]
    return parts if parts else [s]


def split_fio_triple(fio: str) -> tuple[str, str, str]:
    """
    Фамилия, имя, отчество. Все слова после имени объединяются в отчество
    (двойные отчества, «оглы», несколько частей в одной ячейке).
    """
    s = (fio or "").replace("\xa0", " ").strip()
    if not s:
        return ("", "", "")
    parts = [p for p in re.split(r"\s+", s) if p]
    if not parts:
        return ("", "", "")
    if len(parts) == 1:
        return (parts[0], "", "")
    if len(parts) == 2:
        return (parts[0], parts[1].rstrip("."), "")
    fam = parts[0]
    first = parts[1].rstrip(".")
    patronymic = " ".join(parts[2:]).strip()
    return (fam, first, patronymic)


def _norm_fio_lookup_key(fio: str) -> str:
    s = (fio or "").replace("\xa0", " ").strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", s)


def build_snils_position_lookup(employees: list[EmployeeRecord] | None) -> dict[str, tuple[str, str]]:
    """Нормализованное ФИО → (СНИЛС, должность); первая запись при дубликатах."""
    if not employees:
        return {}
    m: dict[str, tuple[str, str]] = {}
    for r in employees:
        k = _norm_fio_lookup_key(r.fio)
        if k and k not in m:
            m[k] = ((r.snils or "").strip(), (r.profession or "").strip())
    return m


def _norm_position_key(s: str) -> str:
    t = (s or "").strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", t)


def employee_enriched_for_v_prof_export(
    emp: EmployeeRecord,
    employees: list[EmployeeRecord] | None,
    lookup: dict[str, tuple[str, str]],
) -> EmployeeRecord:
    """
    В журнале в export_meta_json обычно нет должности: для блока «В» нужны profession/profession2,
    чтобы прочитать лист V_PROF. Берём их из файла сотрудников по ФИО (аналогично СНИЛС в шаблоне).
    """
    fk = _norm_fio_lookup_key(emp.fio or "")
    p1 = (emp.profession or "").strip()
    p2 = (emp.profession2 or "").strip()
    sn = (emp.snils or "").strip()
    sub = (emp.subdivision or "").strip()
    if fk:
        sn_lu, pos_lu = lookup.get(fk, ("", ""))
        pos_lu = (pos_lu or "").strip()
        if not p1 and pos_lu:
            p1 = pos_lu
        if not sn:
            sn = (sn_lu or "").strip()
    if employees and fk:
        ordered: list[str] = []
        seen_norm: set[str] = set()
        for r in employees:
            if _norm_fio_lookup_key(r.fio or "") != fk:
                continue
            for px in ((r.profession or "").strip(), (r.profession2 or "").strip()):
                if not px:
                    continue
                nk = _norm_position_key(px)
                if nk in seen_norm:
                    continue
                seen_norm.add(nk)
                ordered.append(px)
        if not p1 and ordered:
            p1 = ordered[0]
        if not p2 and ordered:
            k1 = _norm_position_key(p1)
            for t in ordered:
                if _norm_position_key(t) != k1:
                    p2 = t
                    break
    if (
        p1 == (emp.profession or "").strip()
        and p2 == (emp.profession2 or "").strip()
        and sn == (emp.snils or "").strip()
    ):
        return emp
    return EmployeeRecord(
        fio=emp.fio,
        profession=p1,
        subdivision=sub,
        profession2=p2,
        snils=sn,
    )


def mintrud_b_program_fragments_for_employee(
    path: Path,
    emp: EmployeeRecord,
    employees: list[EmployeeRecord] | None,
    lookup: dict[str, tuple[str, str]],
) -> list[str]:
    """
    Запасной источник для «Б» в Минтруде, если нет названия с листа B (журнал / кэш).

    По каждой уникальной должности: лист V_PROF, столбец B (якорь), как для разметки таблицы в Word.
    Если строк в файле нет — одна запись с должностью из lookup (как для блока «В»).
    """
    from protocol_docx import (
        _collect_unique_professions_ordered,
        _norm_profession_key,
        _v_prof_anchor_line_from_row,
        _v_prof_select_best_row,
        expand_persons_block_b_rows,
        v_prof_layout_for_path,
    )

    fk = _norm_fio_lookup_key(emp.fio or "")
    rows_src: list[EmployeeRecord] = []
    if employees and fk:
        for r in employees:
            if _norm_fio_lookup_key(r.fio or "") == fk:
                rows_src.append(r)
    if not rows_src:
        rows_src = [employee_enriched_for_v_prof_export(emp, employees, lookup)]

    expanded = expand_persons_block_b_rows(rows_src)
    profs = _collect_unique_professions_ordered(expanded)
    out: list[str] = []
    seen: set[str] = set()
    for pr in profs:
        row = _v_prof_select_best_row(path, pr)
        if not row:
            continue
        layout = v_prof_layout_for_path(path)
        line = _v_prof_anchor_line_from_row(row, layout.col_b_one_based)
        if not line:
            continue
        core = line.lstrip("-").strip()
        if not core:
            continue
        nk = _norm_profession_key(core)
        if nk in seen:
            continue
        seen.add(nk)
        out.append(core)
    return out


def grade_to_test_passed_numeric(grade: str) -> int | None:
    """Колонка «тест пройден»: 1 — удовлетворительно, 0 — неудовлетворительно."""
    g = (grade or "").strip().lower()
    if "неудовлетвор" in g:
        return 0
    if "удовлетвор" in g:
        return 1
    return None


def _parse_export_meta(raw: object) -> dict[str, Any] | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    try:
        d = json.loads(s)
    except (json.JSONDecodeError, TypeError, ValueError):
        return None
    if not isinstance(d, dict):
        return None
    pk = d.get("program_keys")
    pt = d.get("program_titles")
    pr = d.get("persons_raw")
    if pr is None:
        pr = []
    if not isinstance(pk, list) or not isinstance(pt, dict) or not isinstance(pr, list):
        return None
    return d


def _persons_from_meta(persons_raw: list[Any]) -> list[EmployeeRecord]:
    out: list[EmployeeRecord] = []
    for item in persons_raw:
        if not isinstance(item, dict):
            continue
        out.append(
            EmployeeRecord(
                fio=str(item.get("fio", "")),
                profession=str(item.get("profession", "")),
                subdivision=str(item.get("subdivision", "")),
                profession2=str(item.get("profession2", "")),
                snils=str(item.get("snils", "")),
            )
        )
    return out


def _meta_v_prof_context(meta: dict[str, Any]) -> dict[str, Any]:
    """Контекст программы «В» из export_meta_json (снимок протокола)."""
    row_src = _persons_from_meta(meta.get("persons_row_source") or [])
    if not row_src:
        merged = _persons_from_meta(meta.get("persons_raw") or [])
        if merged:
            from protocol_docx import expand_persons_block_b_rows

            row_src = expand_persons_block_b_rows(merged)
        else:
            row_src = []
    face_sheet = str(meta.get("face_sheet_profession") or "").strip() or None
    enabled_by_fio: dict[str, frozenset[str]] | None = None
    enabled_raw = meta.get("v_prof_enabled_by_fio")
    if isinstance(enabled_raw, dict):
        enabled_by_fio = {}
        for k, vals in enabled_raw.items():
            if isinstance(vals, list):
                enabled_by_fio[str(k)] = frozenset(str(x) for x in vals if str(x).strip())
    main_by_fio: dict[str, str] | None = None
    main_raw = meta.get("v_prof_main_by_fio")
    if isinstance(main_raw, dict):
        main_by_fio = {str(k): str(v).strip() for k, v in main_raw.items() if str(v).strip()}
    return {
        "persons_row_source": row_src,
        "face_sheet_profession": face_sheet,
        "v_prof_enabled_by_fio": enabled_by_fio,
        "v_prof_main_by_fio": main_by_fio,
    }


def mintrud_v_program_entries_for_employee(
    catalog_path: Path,
    emp: EmployeeRecord,
    meta: dict[str, Any],
    *,
    employees: list[EmployeeRecord] | None,
    lookup: dict[str, tuple[str, str]],
    v_parts_for_employee: Callable[[Path, EmployeeRecord], list[str]] | None = None,
) -> list[tuple[str, str]]:
    """
    Строки программы «В» для Минтруда: (наименование, должность из протокола).
    Использует снимок протокола в meta; при отсутствии — запасной v_parts_for_employee.
    """
    from protocol_docx import v_program_parts_with_professions_for_employee

    ctx = _meta_v_prof_context(meta)
    if catalog_path.is_file():
        emp_v = employee_enriched_for_v_prof_export(emp, employees, lookup)
        entries = v_program_parts_with_professions_for_employee(
            catalog_path,
            emp_v,
            face_sheet_profession=ctx["face_sheet_profession"],
            persons_row_source=ctx["persons_row_source"],
            v_prof_enabled_by_fio=ctx["v_prof_enabled_by_fio"],
            v_prof_main_by_fio=ctx["v_prof_main_by_fio"],
        )
        if entries:
            return entries
    if v_parts_for_employee is not None and catalog_path.is_file():
        emp_v = employee_enriched_for_v_prof_export(emp, employees, lookup)
        try:
            parts = list(v_parts_for_employee(catalog_path, emp_v) or [])
        except Exception:
            _log.exception(
                "v_parts_for_employee fallback: %s, сотрудник %s",
                catalog_path,
                emp.fio,
            )
            parts = []
        if parts:
            pos = (emp_v.profession or "").strip()
            if not pos:
                _, pos_lu = lookup.get(_norm_fio_lookup_key(emp.fio or ""), ("", ""))
                pos = (pos_lu or "").strip()
            return [(frag, pos) for frag in parts if (frag or "").strip()]
    titles_map = {str(k): str(v) for k, v in meta.get("program_titles", {}).items()}
    fb_parts = v_parts_from_stored_v_title(titles_map.get("V", ""))
    if fb_parts:
        pos = (emp.profession or "").strip()
        if not pos and ctx["persons_row_source"]:
            for row in ctx["persons_row_source"]:
                if _norm_fio_lookup_key(row.fio or "") == _norm_fio_lookup_key(emp.fio or ""):
                    pos = (row.profession or "").strip()
                    if pos:
                        break
        if not pos:
            _, pos_lu = lookup.get(_norm_fio_lookup_key(emp.fio or ""), ("", ""))
            pos = (pos_lu or "").strip()
        return [(frag, pos) for frag in fb_parts]
    return []


def v_parts_from_stored_v_title(title: str) -> list[str]:
    """
    Строки программы «В» из заголовка журнала («Программа (В)\\n(2. …\\n3. …)»).
    Используется, если по V_PROF не удалось получить фрагменты для сотрудника.
    """
    t = (title or "").replace("\xa0", " ").strip()
    if not t:
        return []
    for prefix in ("Программа (В)", "Программа «В»", "Программа обучения «В»"):
        if t.startswith(prefix):
            t = t[len(prefix) :].strip()
            break
    if t.startswith("(") and t.endswith(")"):
        inner = t[1:-1].strip()
        if inner:
            t = inner
    lines = [ln.strip() for ln in t.replace("\r", "").split("\n") if ln.strip()]
    parts: list[str] = []
    for ln in lines:
        m = re.match(r"^(\d+)\s*[\.\)]\s*(.+)$", ln, re.DOTALL)
        parts.append((m.group(2) if m else ln).strip())
    return [p for p in parts if p]


def _make_row(
    *,
    rid: Any,
    ln: str,
    fn: str,
    pt: str,
    snils: str,
    position: str,
    date_s: str,
    proto_no: str,
    tp: int | None,
    inn_s: str,
    name_s: str,
    inn2_s: str,
    org2_s: str,
    program_name: str,
    registry_id: str,
) -> dict[str, Any]:
    row_d: dict[str, Any] = {
        "program_name": program_name,
        "registry_program_id": registry_id,
        "last_name": ln,
        "first_name": fn,
        "patronymic": pt,
        "snils": snils,
        "position": position,
        "doc_date": date_s,
        "protocol_no": proto_no,
        "inn_employer": inn_s,
        "employer_name": name_s,
        "inn_org2": inn2_s,
        "org2_name": org2_s,
        "_journal_id": str(rid),
    }
    if tp is not None:
        row_d["test_passed"] = tp
    return row_d


def build_export_rows(
    records: list[dict[str, Any]],
    *,
    inn_employer: str = "",
    employer_name: str = "",
    inn_org2: str = "",
    org2_name: str = "",
    employees: list[EmployeeRecord] | None = None,
    employees_excel_path: Path | None = None,
    programs_excel_path: Path | None = None,
    v_parts_for_employee: Callable[[Path, EmployeeRecord], list[str]] | None = None,
) -> list[dict[str, Any]]:
    inn_s = (inn_employer or "").strip()
    name_s = (employer_name or "").strip()
    inn2_s = (inn_org2 or "").strip()
    org2_s = (org2_name or "").strip()
    lookup = build_snils_position_lookup(employees)
    emp_excel = (
        Path(employees_excel_path).expanduser().resolve()
        if employees_excel_path is not None
        else None
    )
    prog_excel = (
        Path(programs_excel_path).expanduser().resolve()
        if programs_excel_path is not None
        else None
    )
    catalog_path = (
        prog_excel
        if prog_excel is not None and prog_excel.is_file()
        else emp_excel
    )
    v_rows: list[tuple[str, str, str, float | None]] = []
    if catalog_path is not None and catalog_path.is_file():
        v_rows = get_cached_v_registry_rows(catalog_path)

    out: list[dict[str, Any]] = []
    for r in records:
        rid = r.get("id", "")
        date_s = (r.get("date") or "").strip()
        grade = (r.get("grade") or "").strip()
        content = r.get("content") or ""
        meta = _parse_export_meta(r.get("export_meta_json"))
        proto_no = extract_protocol_number_from_plain_text(str(content))
        if meta:
            pn_m = str(meta.get("protocol_no", "")).strip()
            if pn_m:
                proto_no = pn_m
        tp = grade_to_test_passed_numeric(grade)

        if meta:
            keys = [str(k) for k in meta.get("program_keys", [])]
            titles_map = {str(k): str(v) for k, v in meta.get("program_titles", {}).items()}
            persons = _persons_from_meta(meta.get("persons_raw", []))
            if not persons:
                persons = [
                    EmployeeRecord(fio=x)
                    for x in split_journal_fio_field(str(r.get("fio") or ""))
                    if (x or "").strip()
                ]

            if not keys:
                topic_fb = (r.get("topic") or "").strip()
                for emp in persons:
                    fio_one = (emp.fio or "").strip()
                    if not fio_one:
                        continue
                    ln, fn, pt = split_fio_triple(fio_one)
                    sn_lu, pos_lu = lookup.get(_norm_fio_lookup_key(fio_one), ("", ""))
                    snils = (sn_lu or "").strip()
                    position = (pos_lu or "").strip()
                    out.append(
                        _make_row(
                            rid=rid,
                            ln=ln,
                            fn=fn,
                            pt=pt,
                            snils=snils,
                            position=position,
                            date_s=date_s,
                            proto_no=proto_no,
                            tp=tp,
                            inn_s=inn_s,
                            name_s=name_s,
                            inn2_s=inn2_s,
                            org2_s=org2_s,
                            program_name=topic_fb,
                            registry_id="",
                        )
                    )
                continue

            for emp in persons:
                fio_one = (emp.fio or "").strip()
                if not fio_one:
                    continue
                ln, fn, pt = split_fio_triple(fio_one)
                sn_lu, pos_lu = lookup.get(_norm_fio_lookup_key(fio_one), ("", ""))
                snils = (sn_lu or "").strip()
                position = (pos_lu or "").strip()

                for key in keys:
                    if parse_program_key(key) == ProgramKey.V:
                        entries: list[tuple[str, str]] = []
                        if catalog_path is not None and catalog_path.is_file():
                            try:
                                entries = mintrud_v_program_entries_for_employee(
                                    catalog_path,
                                    emp,
                                    meta,
                                    employees=employees,
                                    lookup=lookup,
                                    v_parts_for_employee=v_parts_for_employee,
                                )
                            except Exception:
                                _log.exception(
                                    "mintrud_v_program_entries_for_employee: %s, сотрудник %s",
                                    catalog_path,
                                    fio_one,
                                )
                                entries = []
                        if not entries:
                            fb_title = titles_map.get(key, "").strip()
                            fb_parts = v_parts_from_stored_v_title(fb_title)
                            if not fb_parts and fb_title:
                                fb_parts = [fb_title]
                            proto_pos = (emp.profession or "").strip()
                            if not proto_pos:
                                for row in _persons_from_meta(
                                    meta.get("persons_row_source") or meta.get("persons_raw") or []
                                ):
                                    if _norm_fio_lookup_key(row.fio or "") == _norm_fio_lookup_key(
                                        fio_one
                                    ):
                                        proto_pos = (row.profession or "").strip()
                                        if proto_pos:
                                            break
                            row_pos_fb = proto_pos or position
                            for frag in fb_parts:
                                frag = (frag or "").strip()
                                if not frag:
                                    continue
                                gid = _registry_id_from_v(frag, v_rows)
                                out.append(
                                    _make_row(
                                        rid=rid,
                                        ln=ln,
                                        fn=fn,
                                        pt=pt,
                                        snils=snils,
                                        position=row_pos_fb,
                                        date_s=date_s,
                                        proto_no=proto_no,
                                        tp=tp,
                                        inn_s=inn_s,
                                        name_s=name_s,
                                        inn2_s=inn2_s,
                                        org2_s=org2_s,
                                        program_name=frag,
                                        registry_id=gid,
                                    )
                                )
                            continue
                        for frag, v_position in entries:
                            frag = (frag or "").strip()
                            if not frag:
                                continue
                            row_position = (v_position or "").strip() or position
                            gid = _registry_id_from_v(frag, v_rows)
                            out.append(
                                _make_row(
                                    rid=rid,
                                    ln=ln,
                                    fn=fn,
                                    pt=pt,
                                    snils=snils,
                                    position=row_position,
                                    date_s=date_s,
                                    proto_no=proto_no,
                                    tp=tp,
                                    inn_s=inn_s,
                                    name_s=name_s,
                                    inn2_s=inn2_s,
                                    org2_s=org2_s,
                                    program_name=frag,
                                    registry_id=gid,
                                )
                            )
                    elif parse_program_key(key) == ProgramKey.B:
                        raw_title = ""
                        if catalog_path is not None and catalog_path.is_file():
                            raw_title = get_cached_b_program_title(catalog_path).strip()
                        if not raw_title:
                            raw_title = titles_map.get(key, "").strip()
                        if not raw_title:
                            continue
                        pname, gid = _registry_title_and_id_from_v(raw_title, v_rows)
                        if not pname:
                            pname = raw_title
                        out.append(
                            _make_row(
                                rid=rid,
                                ln=ln,
                                fn=fn,
                                pt=pt,
                                snils=snils,
                                position=position,
                                date_s=date_s,
                                proto_no=proto_no,
                                tp=tp,
                                inn_s=inn_s,
                                name_s=name_s,
                                inn2_s=inn2_s,
                                org2_s=org2_s,
                                program_name=pname,
                                registry_id=gid,
                            )
                        )
                    elif parse_program_key(key) == ProgramKey.PP:
                        raw_title = ""
                        if catalog_path is not None and catalog_path.is_file():
                            raw_title = get_cached_pp_table_title(catalog_path).strip()
                        if not raw_title:
                            raw_title = titles_map.get(key, "").strip()
                        if not raw_title:
                            continue
                        pname, gid = _registry_title_and_id_from_v(raw_title, v_rows)
                        if not pname:
                            pname = raw_title
                        out.append(
                            _make_row(
                                rid=rid,
                                ln=ln,
                                fn=fn,
                                pt=pt,
                                snils=snils,
                                position=position,
                                date_s=date_s,
                                proto_no=proto_no,
                                tp=tp,
                                inn_s=inn_s,
                                name_s=name_s,
                                inn2_s=inn2_s,
                                org2_s=org2_s,
                                program_name=pname,
                                registry_id=gid,
                            )
                        )
                    elif parse_program_key(key) == ProgramKey.SIZ:
                        raw_title = ""
                        if catalog_path is not None and catalog_path.is_file():
                            raw_title = get_cached_siz_table_title(catalog_path).strip()
                        if not raw_title:
                            raw_title = titles_map.get(key, "").strip()
                        if not raw_title:
                            continue
                        pname, gid = _registry_title_and_id_from_v(raw_title, v_rows)
                        if not pname:
                            pname = raw_title
                        out.append(
                            _make_row(
                                rid=rid,
                                ln=ln,
                                fn=fn,
                                pt=pt,
                                snils=snils,
                                position=position,
                                date_s=date_s,
                                proto_no=proto_no,
                                tp=tp,
                                inn_s=inn_s,
                                name_s=name_s,
                                inn2_s=inn2_s,
                                org2_s=org2_s,
                                program_name=pname,
                                registry_id=gid,
                            )
                        )
                    else:
                        raw_title = titles_map.get(key, "").strip()
                        if not raw_title:
                            continue
                        pname, gid = _registry_title_and_id_from_v(raw_title, v_rows)
                        if not pname:
                            pname = raw_title
                        out.append(
                            _make_row(
                                rid=rid,
                                ln=ln,
                                fn=fn,
                                pt=pt,
                                snils=snils,
                                position=position,
                                date_s=date_s,
                                proto_no=proto_no,
                                tp=tp,
                                inn_s=inn_s,
                                name_s=name_s,
                                inn2_s=inn2_s,
                                org2_s=org2_s,
                                program_name=pname,
                                registry_id=gid,
                            )
                        )
            continue

        topic = (r.get("topic") or "").strip()
        for fio_one in split_journal_fio_field(str(r.get("fio") or "")):
            ln, fn, pt = split_fio_triple(fio_one)
            snils, position = lookup.get(_norm_fio_lookup_key(fio_one), ("", ""))
            out.append(
                _make_row(
                    rid=rid,
                    ln=ln,
                    fn=fn,
                    pt=pt,
                    snils=snils,
                    position=position,
                    date_s=date_s,
                    proto_no=proto_no,
                    tp=tp,
                    inn_s=inn_s,
                    name_s=name_s,
                    inn2_s=inn2_s,
                    org2_s=org2_s,
                    program_name=topic,
                    registry_id="",
                )
            )
    return out


def export_row_dedupe_key(row: dict[str, Any]) -> tuple[str, ...]:
    """Ключ строки выгрузки Минтруд для слияния без дублей."""
    return (
        (row.get("protocol_no") or "").strip().casefold(),
        format_mintrud_iso_date(row.get("doc_date")),
        (row.get("last_name") or "").strip().casefold(),
        (row.get("first_name") or "").strip().casefold(),
        (row.get("patronymic") or "").strip().casefold(),
        (row.get("program_name") or "").strip().casefold(),
    )


def dedupe_export_data_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Убрать дубликаты строк (последняя побеждает)."""
    merged: dict[tuple[str, ...], dict[str, Any]] = {}
    for row in rows:
        merged[export_row_dedupe_key(row)] = row
    return list(merged.values())


def merge_export_data_rows(
    existing: list[dict[str, Any]],
    new_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Объединить выгрузку: новые строки перезаписывают совпадающие по ключу.

    Сейчас не используется при сохранении Excel (выгрузка только из выбранных записей журнала);
    оставлено для тестов и возможного режима «дополнить файл» в будущем.
    """
    merged: dict[tuple[str, ...], dict[str, Any]] = {
        export_row_dedupe_key(r): dict(r) for r in existing
    }
    for row in new_rows:
        merged[export_row_dedupe_key(row)] = dict(row)
    return list(merged.values())


def read_data_rows_from_mintrud_sheet(ws: Any, field_col: dict[str, int]) -> list[dict[str, Any]]:
    """Прочитать уже заполненные строки из сохранённого шаблона Минтруд."""
    rows: list[dict[str, Any]] = []
    if not field_col:
        return rows
    max_row = int(ws.max_row or HEADER_ROW)
    for r in range(HEADER_ROW + 1, max_row + 1):
        row_data: dict[str, Any] = {}
        has_data = False
        for field, col in field_col.items():
            val = ws.cell(row=r, column=col).value
            if val is None or (isinstance(val, str) and not val.strip()):
                row_data[field] = None
                continue
            has_data = True
            if field == "doc_date":
                row_data[field] = format_mintrud_iso_date(val)
            elif field == "test_passed":
                if isinstance(val, bool):
                    row_data[field] = 1 if val else 0
                elif isinstance(val, (int, float)):
                    row_data[field] = int(val)
                else:
                    s = str(val).strip().lower()
                    row_data[field] = 1 if s in {"1", "да", "true", "yes"} else 0
            else:
                row_data[field] = str(val).strip() if not isinstance(val, str) else val.strip()
        if has_data:
            rows.append(row_data)
    return rows


def _normalize_header_cell(value: object) -> str:
    s = str(value or "").strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", s)


def _header_column_map(ws: Any, header_row: int) -> dict[str, int]:
    m: dict[str, int] = {}
    max_c = int(ws.max_column or 0)
    last_col = max(max_c, 40)
    for c in range(1, last_col + 1):
        v = ws.cell(row=header_row, column=c).value
        key = _normalize_header_cell(v)
        if key:
            m[key] = c
    return m


def _field_to_column(header_cols: dict[str, int]) -> dict[str, int]:
    out: dict[str, int] = {}
    for h_norm, field in FIELD_BY_HEADER_NORM.items():
        col = header_cols.get(h_norm)
        if col is not None:
            out[field] = col
    return out


def _xml_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value == int(value) else str(value)
    return str(value).strip()


def _xml_sub_element(parent: ET.Element, tag: str, value: object) -> ET.Element:
    el = ET.SubElement(parent, tag)
    text = _xml_text(value)
    if text:
        el.text = text
    return el


def _test_passed_attr(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        return str(int(value))
    s = str(value).strip()
    return s


def write_mintrud_export_xml(path: Path, data_rows: list[dict[str, Any]]) -> None:
    """
    XML RegistrySet по XSD шаблона Минтруда (для загрузки на сайт и проверки).
    Структура совпадает с «Шаблон_Минтруд_XSD_УМН как должно быть.xml».
    """
    root = ET.Element("RegistrySet")
    for rd in data_rows:
        rec = ET.SubElement(root, "RegistryRecord")
        worker = ET.SubElement(rec, "Worker")
        _xml_sub_element(worker, "LastName", rd.get("last_name"))
        _xml_sub_element(worker, "FirstName", rd.get("first_name"))
        _xml_sub_element(worker, "MiddleName", rd.get("patronymic"))
        _xml_sub_element(worker, "Position", rd.get("position"))
        _xml_sub_element(worker, "Snils", rd.get("snils"))
        _xml_sub_element(worker, "EmployerInn", rd.get("inn_employer"))
        _xml_sub_element(worker, "EmployerTitle", rd.get("employer_name"))
        org = ET.SubElement(rec, "Organization")
        _xml_sub_element(org, "Inn", rd.get("inn_org2"))
        _xml_sub_element(org, "Title", rd.get("org2_name"))
        test = ET.SubElement(rec, "Test")
        test.set("isPassed", _test_passed_attr(rd.get("test_passed")))
        test.set("learnProgramId", _xml_text(rd.get("registry_program_id")))
        _xml_sub_element(
            test, "Date", format_mintrud_iso_date(rd.get("doc_date"))
        )
        _xml_sub_element(test, "ProtocolNumber", rd.get("protocol_no"))
        _xml_sub_element(test, "LearnProgramTitle", rd.get("program_name"))
    try:
        ET.indent(root, space="\t")
    except (AttributeError, TypeError):
        pass
    out = Path(path).expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    body = ET.tostring(root, encoding="unicode")
    with out.open("w", encoding="utf-8", newline="\n") as f:
        f.write('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n')
        f.write(body)
        if not body.endswith("\n"):
            f.write("\n")


_MINTRUD_OPENPYXL_DATA_PARTS: frozenset[str] = frozenset(
    {
        "xl/worksheets/sheet1.xml",
        "xl/tables/table1.xml",
        "xl/styles.xml",
    }
)


def _fix_sheet1_table_rel_id(sheet_xml: bytes) -> bytes:
    """
    В шаблоне XSD таблица на листе связана как rId2 (см. sheet1.xml.rels);
    openpyxl пишет rId1 — Excel считает книгу повреждённой.
    """
    text = sheet_xml.decode("utf-8")
    text = re.sub(
        r"(<tablePart\b[^>]*\br:id=\")rId\d+(\")",
        r"\1rId2\2",
        text,
        count=1,
    )
    return text.encode("utf-8")


def _merge_mintrud_workbook_from_template(
    template_path: Path, edited_path: Path
) -> None:
    """
    openpyxl save() портит XSD-каркас (xmlMaps, rels, Content_Types).
    Собираем итог: структура пакета из шаблона, данные листа/таблицы — из openpyxl.
    """
    template_path = Path(template_path).resolve()
    edited_path = Path(edited_path).resolve()
    if not template_path.is_file() or not edited_path.is_file():
        return
    with zipfile.ZipFile(template_path, "r") as zt:
        if "xl/xmlMaps.xml" not in zt.namelist():
            return

    data_parts: dict[str, bytes] = {}
    with zipfile.ZipFile(edited_path, "r") as zd:
        for info in zd.infolist():
            data_parts[info.filename] = zd.read(info.filename)

    tmp_path = edited_path.with_suffix(edited_path.suffix + ".__merge")
    try:
        with zipfile.ZipFile(template_path, "r") as zt:
            with zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as zw:
                for info in zt.infolist():
                    name = info.filename
                    if name == "xl/worksheets/sheet1.xml" and name in data_parts:
                        payload = _fix_sheet1_table_rel_id(data_parts[name])
                    elif name in _MINTRUD_OPENPYXL_DATA_PARTS and name in data_parts:
                        payload = data_parts[name]
                    else:
                        payload = zt.read(name)
                    zw.writestr(info, payload)
        shutil.move(str(tmp_path), str(edited_path))
    finally:
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)


def _extend_mintrud_sheet_table(
    ws: Any,
    *,
    header_row: int,
    last_data_row: int,
    last_col: int = 14,
) -> None:
    """
    Расширяет XML-таблицу листа «Шаблон» (по умолчанию A1:N4) до последней строки данных.
    Иначе Excel/сайт видят только 3 строки шаблона — «пустые» строки при загрузке.
    """
    if last_data_row <= header_row:
        return
    tables = getattr(ws, "_tables", None)
    if not tables:
        return
    try:
        from openpyxl.utils import get_column_letter
    except ImportError:
        return
    max_col = max(int(last_col), 14)
    new_ref = f"A1:{get_column_letter(max_col)}{last_data_row}"
    for tbl in tables.values():
        tbl.ref = new_ref
        if getattr(tbl, "autoFilter", None) is not None:
            tbl.autoFilter.ref = new_ref


def _resolve_employees_list(
    employees: list[EmployeeRecord] | None,
    employees_excel_path: Path | None,
) -> list[EmployeeRecord]:
    if employees is not None:
        return employees
    if employees_excel_path is None or not employees_excel_path.is_file():
        return []
    try:
        return load_employees_from_excel(employees_excel_path)
    except EmployeeExcelError:
        return []


def write_mintrud_template_xlsx(
    path: Path,
    records: list[dict[str, Any]],
    *,
    template_path: Path | None = None,
    inn_employer: str = "",
    employer_name: str = "",
    inn_org2: str = "",
    org2_name: str = "",
    employees: list[EmployeeRecord] | None = None,
    employees_excel_path: Path | None = None,
    programs_excel_path: Path | None = None,
    v_parts_for_employee: Callable[[Path, EmployeeRecord], list[str]] | None = None,
) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("Нужен пакет openpyxl (pip install openpyxl).") from e

    tpl = template_path if template_path is not None else mintrud_template_path()
    if tpl is None or not Path(tpl).is_file():
        names = " или ".join(f"«{n}»" for n in MINTRUD_TEMPLATE_FILENAMES)
        raise FileNotFoundError(
            f"Не найден XSD-шаблон Минтруда. Положите в папку с программой файл {names}."
        )
    tpl = Path(tpl).expanduser().resolve()

    emp_resolved = _resolve_employees_list(employees, employees_excel_path)

    wb = load_workbook(tpl)
    if MINTRUD_TEMPLATE_SHEET in wb.sheetnames:
        ws = wb[MINTRUD_TEMPLATE_SHEET]
    else:
        ws = wb.active

    header_cols = _header_column_map(ws, HEADER_ROW)
    field_col = _field_to_column(header_cols)
    if not field_col:
        raise ValueError(
            "В первой строке листа шаблона не найдены ожидаемые заголовки "
            "(«Фамилия», «Номер протокола» и т.д.). Проверьте файл шаблона."
        )

    data_rows = build_export_rows(
        records,
        inn_employer=inn_employer,
        employer_name=employer_name,
        inn_org2=inn_org2,
        org2_name=org2_name,
        employees=emp_resolved,
        employees_excel_path=employees_excel_path,
        programs_excel_path=programs_excel_path,
        v_parts_for_employee=v_parts_for_employee,
    )
    data_rows = dedupe_export_data_rows(data_rows)

    out_path = Path(path).expanduser().resolve()

    scan_cols = max(int(ws.max_column or 0), 40)
    if ws.max_row > HEADER_ROW:
        for r in range(HEADER_ROW + 1, ws.max_row + 1):
            for c in range(1, scan_cols + 1):
                ws.cell(row=r, column=c, value=None)

    for i, row_data in enumerate(data_rows):
        r = HEADER_ROW + 1 + i
        for field, col in field_col.items():
            if field not in row_data:
                ws.cell(row=r, column=col, value=None)
                continue
            val = row_data[field]
            if field == "doc_date":
                d = parse_mintrud_iso_date(val)
                ws.cell(row=r, column=col, value=d if d is not None else None)
                continue
            if val is None:
                ws.cell(row=r, column=col, value=None)
            elif isinstance(val, str):
                s = val.strip()
                ws.cell(row=r, column=col, value=s if s else None)
            else:
                ws.cell(row=r, column=col, value=val)

    last_row = HEADER_ROW + len(data_rows) if data_rows else HEADER_ROW + 1
    last_col = max(field_col.values()) if field_col else 14
    _extend_mintrud_sheet_table(
        ws, header_row=HEADER_ROW, last_data_row=last_row, last_col=last_col
    )

    wb.save(out_path)
    _merge_mintrud_workbook_from_template(tpl, out_path)
    xml_path = out_path.with_suffix(".xml")
    try:
        write_mintrud_export_xml(xml_path, data_rows)
    except OSError as e:
        raise RuntimeError(
            f"Excel сохранён:\n{out_path}\n\nНе удалось записать XML:\n{e}"
        ) from e
