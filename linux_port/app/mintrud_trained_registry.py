# -*- coding: utf-8 -*-
"""
Выгрузка «Реестр обученных по охране труда лиц» с портала Минтруда (Excel):
поиск номеров записей для подстановки в протокол. Типичный лист: «Номер в реестре»,
«Фамилия» / «Имя» / «Отчество», «СНИЛС», «Номер протокола», «Программа обучения».

Сопоставление строк реестра с сотрудником в протоколе (все этапы обязательны, где данные есть):
1) СНИЛС (только цифры), если указан и в файле, и у сотрудника;
2) иначе нормализованное ФИО;
3) номер протокола — отбор по колонке «номер протокола» (гибкое совпадение);
4) должность — только для блока программы «Б»: при указанной должности сотрудника в реестре должна
   совпасть колонка «Должность»; иначе № из файла не подставляется. Для ПП, СИЗ и «В» фильтр по должности
   не применяется.

Номера для ячеек «результат»:
- в каждый блок попадают только строки реестра, у которых «Программа обучения» согласуется с программой этого блока;
- блок «В»: позиции 1., 2., … в ячейке результата строго по порядку строк в шапке (V_PROF), без подстановки «чужой» строки реестра;
- Б / PP / СИЗ: по заголовку блока; несколько слотов — разные строки реестра с той же программой.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from employees_io import EmployeeRecord
from v_program_registry_match import fg_line_comparison_key


def _norm_header_cell(value: object) -> str:
    s = str(value or "").strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", s)


def _norm_fio_key(fio: str) -> str:
    s = (fio or "").replace("\xa0", " ").strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", s)


def _snils_digits(s: str) -> str:
    return re.sub(r"\D", "", s or "")


def _norm_protocol_cell(s: str) -> str:
    t = (s or "").replace("\xa0", " ").lower()
    t = re.sub(r"[№#]", "", t)
    return re.sub(r"\s+", " ", t).strip()


def _protocol_tokens_for_match(s: str) -> set[str]:
    """Фрагменты для сравнения номеров протокола (полная строка + крупные группы цифр)."""
    n = _norm_protocol_cell(s)
    if not n:
        return set()
    out = {n}
    for m in re.finditer(r"\d{2,}", n):
        out.add(m.group(0))
    return out


def _protocol_row_matches_queries(row_proto: str, queries: list[str]) -> bool:
    qtok: set[str] = set()
    for q in queries:
        if not (q or "").strip():
            continue
        qtok |= _protocol_tokens_for_match(q)
    if not qtok:
        return True
    rtok = _protocol_tokens_for_match(row_proto)
    if not rtok:
        return True
    rn = _norm_protocol_cell(row_proto)
    qn = [_norm_protocol_cell(q) for q in queries if (q or "").strip()]
    for q in qn:
        if q and (q in rn or rn in q):
            return True
    return bool(qtok & rtok)


HEADER_FIELD_PATTERNS: dict[str, tuple[str, ...]] = {
    "fio": ("фио", "работник", "слушатель"),
    "snils": ("снилс", "страховой номер", "индивидуального лицевого"),
    "protocol": ("номер протокола", "протокол проверки", "№ протокола"),
    "program": (
        "наименование программы",
        "программа обучения",
        "программа по охране",
        "программа",
    ),
    "registry": (
        "номер в реестре",
        "регистрационный номер",
        "номер записи",
        "id записи",
        "реестровый",
        "№ в реестре",
        "уникальный номер",
        "идентификатор записи",
        "номер регистрации",
    ),
    "doc_date": ("дата проверки", "дата проверки знаний", "дата в удостоверении"),
    "position": (
        "должность",
        "занимаемая должность",
        "профессия",
        "должность работника",
    ),
}


def _header_field_for_text(h: str) -> str | None:
    for field_key, pats in HEADER_FIELD_PATTERNS.items():
        for p in pats:
            if p in h:
                return field_key
    return None


def _split_fio_header_field(h: str) -> str | None:
    """Точное совпадение заголовков выгрузки Минтруда: Фамилия / Имя / Отчество."""
    if h == "фамилия":
        return "lastname"
    if h == "имя":
        return "firstname"
    if h == "отчество":
        return "patronymic"
    return None


def _detect_header_map(ws: Any, max_scan_row: int = 12) -> tuple[int, dict[str, int]]:
    """(номер_строки_заголовка, поле -> индекс_колонки_1_based)."""
    max_r = min(int(ws.max_row or 0), max_scan_row)
    max_c = min(int(ws.max_column or 0), 60)
    if max_r < 1 or max_c < 1:
        return 1, {}
    best_row = 1
    best_map: dict[str, int] = {}
    best_score = 0
    for r in range(1, max_r + 1):
        col_map: dict[str, int] = {}
        for c in range(1, max_c + 1):
            h = _norm_header_cell(ws.cell(row=r, column=c).value)
            if not h:
                continue
            role = _split_fio_header_field(h)
            if role:
                if role not in col_map:
                    col_map[role] = c
                continue
            fld = _header_field_for_text(h)
            if fld and fld not in col_map:
                col_map[fld] = c
        score = len(col_map)
        if "fio" in col_map or (
            "lastname" in col_map and "firstname" in col_map
        ):
            score += 2
        if "registry" in col_map:
            score += 2
        if score > best_score:
            best_score = score
            best_row = r
            best_map = col_map
    return best_row, best_map


def _cell_str(ws: Any, r: int, c: int) -> str:
    v = ws.cell(row=r, column=c).value
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d.%m.%Y %H:%M:%S").replace(" 00:00:00", "").strip()
    if isinstance(v, date):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).replace("\xa0", " ").strip()


def _build_row_fio(ws: Any, r: int, col_map: dict[str, int]) -> str:
    if "fio" in col_map:
        return _cell_str(ws, r, col_map["fio"])
    ln = _cell_str(ws, r, col_map["lastname"])
    fn = _cell_str(ws, r, col_map["firstname"])
    pat = (
        _cell_str(ws, r, col_map["patronymic"])
        if "patronymic" in col_map
        else ""
    )
    parts = [x for x in (ln, fn, pat) if x]
    return " ".join(parts)


def _score_program_match(hint: str, program_col: str) -> int:
    a = fg_line_comparison_key(hint)
    b = fg_line_comparison_key(program_col)
    if not a or not b:
        return 0
    if a == b:
        return 4
    if len(b) >= 4 and (a in b or b in a):
        return 3
    if a in b or b in a:
        return 2
    return 0


def _score_profession_match(emp_profession: str, row_position: str) -> int:
    """Совпадение должности сотрудника с колонкой «Должность» строки реестра (как у программ)."""
    return _score_program_match(emp_profession, row_position)


def _narrow_rows_by_employee_profession(
    rows: list[TrainedRegistryRow],
    emp: EmployeeRecord,
) -> list[TrainedRegistryRow]:
    """
    Только для программы «Б»: если у сотрудника указана должность — остаются строки реестра
    с совпадающей «Должностью». При отсутствии совпадений — пустой список (№ из файла не подставляется).
    Если должность сотрудника не заполнена — список не сужается по этому признаку.
    """
    prof = (emp.profession or "").strip()
    if not prof or not rows:
        return rows
    return [r for r in rows if _score_profession_match(prof, r.position) > 0]


@dataclass
class TrainedRegistryRow:
    fio: str
    snils_digits: str
    protocol: str
    program: str
    registry_num: str
    doc_date: str
    position: str = ""


def _build_registry_lookup_indexes(
    rows: list[TrainedRegistryRow],
) -> tuple[dict[str, list[TrainedRegistryRow]], dict[str, list[TrainedRegistryRow]]]:
    by_snils: dict[str, list[TrainedRegistryRow]] = {}
    by_fio: dict[str, list[TrainedRegistryRow]] = {}
    for r in rows:
        if r.snils_digits:
            by_snils.setdefault(r.snils_digits, []).append(r)
        fk = _norm_fio_key(r.fio)
        if fk:
            by_fio.setdefault(fk, []).append(r)
    return by_snils, by_fio


@dataclass
class TrainedRegistryIndex:
    rows: list[TrainedRegistryRow]
    source_path: Path
    _by_snils: dict[str, list[TrainedRegistryRow]] = field(
        default_factory=dict, repr=False
    )
    _by_fio: dict[str, list[TrainedRegistryRow]] = field(
        default_factory=dict, repr=False
    )

    def candidates_for_employee(
        self,
        emp: EmployeeRecord,
        protocol_queries: list[str],
        *,
        apply_profession_filter_for_registry: bool = False,
    ) -> list[TrainedRegistryRow]:
        emp_sn = _snils_digits(emp.snils or "")
        emp_fio = _norm_fio_key(emp.fio or "")

        def _after_protocol(out: list[TrainedRegistryRow]) -> list[TrainedRegistryRow]:
            if apply_profession_filter_for_registry:
                return _narrow_rows_by_employee_profession(out, emp)
            return out

        if self._by_snils or self._by_fio:
            if emp_sn:
                pool = self._by_snils.get(emp_sn, [])
                out = [
                    r
                    for r in pool
                    if _protocol_row_matches_queries(r.protocol, protocol_queries)
                ]
                if out:
                    return _after_protocol(out)
            pool = self._by_fio.get(emp_fio, []) if emp_fio else []
            out = [
                r
                for r in pool
                if _protocol_row_matches_queries(r.protocol, protocol_queries)
            ]
            return _after_protocol(out)
        out: list[TrainedRegistryRow] = []
        if emp_sn:
            for r in self.rows:
                if r.snils_digits and emp_sn == r.snils_digits:
                    if _protocol_row_matches_queries(r.protocol, protocol_queries):
                        out.append(r)
            if out:
                return _after_protocol(out)
        for r in self.rows:
            if emp_fio and _norm_fio_key(r.fio) == emp_fio:
                if _protocol_row_matches_queries(r.protocol, protocol_queries):
                    out.append(r)
        return _after_protocol(out)

    def registry_numbers_for_hints(
        self,
        candidates: list[TrainedRegistryRow],
        hints: list[str],
        emp: EmployeeRecord | None = None,
        *,
        require_profession_for_registry: bool = False,
    ) -> list[str]:
        """
        Для каждой подсказки по порядку (строка шапки блока «В» 1, 2, … или слоты Б/ПП/СИЗ)
        выбирается лучшая неиспользованная строка реестра с ненулевым совпадением программы.
        Для блока «Б» при require_profession_for_registry учитываются только строки с совпадением
        «Должность» в реестре; для ПП/СИЗ/«В» должность лишь различает равные совпадения по программе.
        """
        remaining = list(candidates)
        result: list[str] = []
        emp_prof = (emp.profession or "").strip() if emp is not None else ""
        for hint in hints:
            hint = (hint or "").strip()
            best_i = -1
            best_key = (-1, -1)
            for i, row in enumerate(remaining):
                sc = _score_program_match(hint, row.program) if hint else 0
                if sc <= 0:
                    continue
                pb = _score_profession_match(emp_prof, row.position) if emp_prof else 0
                if emp_prof and require_profession_for_registry and pb <= 0:
                    continue
                key = (sc, pb)
                if key > best_key:
                    best_key = key
                    best_i = i
            if best_i >= 0 and best_key[0] > 0:
                result.append((remaining.pop(best_i).registry_num or "").strip())
            else:
                result.append("")
        return result


def filter_candidates_for_program_block(
    candidates: list[TrainedRegistryRow],
    block_hints: list[str],
) -> list[TrainedRegistryRow]:
    """
    Оставляет строки реестра, у которых «Программа обучения» согласуется хотя бы с одной
    подсказкой блока (заголовок Б/ПП/СИЗ или фрагмент V_PROF). Иначе пусто — в блок не подставляем
    номера из файла (останется ручное поле).
    """
    hints = [h.strip() for h in block_hints if (h or "").strip()]
    if not hints:
        return list(candidates)
    out: list[TrainedRegistryRow] = []
    for row in candidates:
        best = 0
        for h in hints:
            best = max(best, _score_program_match(h, row.program))
        if best > 0:
            out.append(row)
    return out


def load_trained_registry_index(path: Path | None) -> TrainedRegistryIndex | None:
    """Читает первый лист .xlsx; при ошибке или отсутствии колонок — None."""
    if path is None:
        return None
    p = Path(path).expanduser().resolve()
    if not p.is_file():
        return None
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        wb = load_workbook(p, read_only=False, data_only=True)
    except Exception:
        return None
    try:
        ws = wb.active
        header_row, col_map = _detect_header_map(ws)
        has_fio = "fio" in col_map or (
            "lastname" in col_map and "firstname" in col_map
        )
        if "registry" not in col_map or not has_fio:
            return None
        c_reg = col_map["registry"]
        c_snils = col_map.get("snils")
        c_prot = col_map.get("protocol")
        c_prog = col_map.get("program")
        c_date = col_map.get("doc_date")
        c_pos = col_map.get("position")
        rows: list[TrainedRegistryRow] = []
        for r in range(header_row + 1, (ws.max_row or 0) + 1):
            fio = _build_row_fio(ws, r, col_map)
            reg = _cell_str(ws, r, c_reg)
            if not fio and not reg:
                continue
            sn = _snils_digits(_cell_str(ws, r, c_snils)) if c_snils else ""
            prot = _cell_str(ws, r, c_prot) if c_prot else ""
            prog = _cell_str(ws, r, c_prog) if c_prog else ""
            d = _cell_str(ws, r, c_date) if c_date else ""
            pos = _cell_str(ws, r, c_pos) if c_pos else ""
            if not reg:
                continue
            rows.append(
                TrainedRegistryRow(
                    fio=fio,
                    snils_digits=sn,
                    protocol=prot,
                    program=prog,
                    registry_num=reg,
                    doc_date=d,
                    position=pos,
                )
            )
        if not rows:
            return None
        by_snils, by_fio = _build_registry_lookup_indexes(rows)
        return TrainedRegistryIndex(
            rows=rows, source_path=p, _by_snils=by_snils, _by_fio=by_fio
        )
    finally:
        wb.close()


def merge_registry_tokens(
    manual_raw: str,
    n: int,
    file_tokens: list[str] | None,
) -> list[str]:
    """n слотов: сначала непустые из файла, затем из ручного поля (разделители , ; перевод строки)."""
    manual_parts = [
        p.strip()
        for p in re.split(r"[\n,;]+", (manual_raw or "").strip())
        if p.strip()
    ]
    out: list[str] = []
    for i in range(n):
        ft = ""
        if file_tokens and i < len(file_tokens):
            ft = (file_tokens[i] or "").strip()
        if ft:
            out.append(ft)
        elif i < len(manual_parts):
            out.append(manual_parts[i])
        else:
            out.append("")
    return out
