# -*- coding: utf-8 -*-
"""
Кэш в SQLite (protocols.db): справочники программ (листы B, PP, SIZ, V, снимки столбцов V_PROF).

Снимок списка сотрудников в БД не ведётся (персональные данные); список всегда из Excel.
При загрузке сотрудников старые строки кэша сотрудников для этого файла удаляются.

Программы: листы B/PP/SIZ/V и снимки V_PROF (якоря PP/СИЗ в тексте); кэш строк V_PROF очищается при смене файла.
"""

from __future__ import annotations

import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from app_paths import application_user_dir
from employees_io import EmployeeRecord
from program_keys import ProgramKey

DATABASE_FILENAME = "protocols.db"
EXCEL_SYNC_ROLE_EMPLOYEES = "employees"
EXCEL_SYNC_ROLE_PROGRAMS = "programs"

# Соответствует main.py
B_PROGRAM_SHEET_NAME = "B"
B_PROGRAM_TITLE_COL = 2
# Необязательный объём (академические часы) в строке с наименованием программы.
B_PROGRAM_HOURS_COL = 3
# Заголовок блока PP/СИЗ в таблице протокола — первая непустая ячейка столбца 2 (как у «Б»).
PP_TABLE_SHEET_NAME = "PP"
SIZ_TABLE_SHEET_NAME = "SIZ"
V_PROF_SHEET_NAME = "V_PROF"
V_PROF_TITLE_COL_PP = 3
V_PROF_TITLE_COL_SIZ = 4
V_PROGRAM_REGISTRY_SHEET = "V"
V_PROGRAM_SHEET_NAME_ALIASES: tuple[str, ...] = ("v", "в")
V_PROGRAM_GOS_REGISTRY_ID_COL_A = 1
V_PROGRAM_MATCH_COL_B = 2
V_PROGRAM_TITLE_COL_C = 3
# Объём программы «В» (ч) на листе V — в той же строке, что и данные реестра (столбец B не пуст), столбец D.
V_PROGRAM_HOURS_COL_D = 4
V_PROGRAM_SHEET_MAX_ROWS = 2000

# Увеличивайте при изменении логики кэша программ, чтобы перечитать Excel без смены mtime файла.
PROGRAM_CATALOG_CACHE_SCHEMA = 5

_MTIME_EPS = 1e-5


def parse_training_hours_value(val: object) -> float | None:
    """Число часов из ячейки Excel: число или текст вроде «40», «40,5», «40 ч»."""
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and (val != val or val in (float("inf"), float("-inf"))):
            return None
        return float(val)
    s0 = str(val).strip()
    if not s0:
        return None
    s = s0.replace("\xa0", " ").replace(",", ".")
    m = re.search(r"[-+]?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def _db_path() -> Path:
    return application_user_dir() / DATABASE_FILENAME


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def ensure_excel_cache_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS excel_sync_state (
            file_role TEXT PRIMARY KEY,
            path TEXT NOT NULL,
            mtime REAL NOT NULL,
            size_bytes INTEGER NOT NULL,
            synced_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS cache_b_program_title (
            file_path TEXT NOT NULL,
            file_mtime REAL NOT NULL,
            title_text TEXT NOT NULL,
            source_row INTEGER,
            hours_value REAL,
            PRIMARY KEY (file_path)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS cache_v_registry (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_path TEXT NOT NULL,
            file_mtime REAL NOT NULL,
            row_order INTEGER NOT NULL,
            gos_id TEXT NOT NULL,
            col_b TEXT NOT NULL,
            col_c TEXT NOT NULL,
            col_d_hours REAL,
            UNIQUE (file_path, row_order)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS cache_v_prof_column_snapshot (
            file_path TEXT NOT NULL,
            file_mtime REAL NOT NULL,
            column_key TEXT NOT NULL,
            first_nonempty TEXT NOT NULL,
            PRIMARY KEY (file_path, column_key)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS cache_pp_siz_table_header (
            file_path TEXT NOT NULL,
            file_mtime REAL NOT NULL,
            pp_title TEXT NOT NULL,
            siz_title TEXT NOT NULL,
            pp_hours REAL,
            siz_hours REAL,
            v_prof_hours REAL,
            cache_schema INTEGER NOT NULL DEFAULT 2,
            PRIMARY KEY (file_path)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS cache_employees_snapshot (
            file_path TEXT PRIMARY KEY,
            file_mtime REAL NOT NULL,
            size_bytes INTEGER NOT NULL,
            employees_json TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS v_prof_cache (
            file_path TEXT NOT NULL,
            profession_norm TEXT NOT NULL,
            file_mtime REAL NOT NULL,
            row_text TEXT NOT NULL,
            parts_json TEXT,
            PRIMARY KEY (file_path, profession_norm)
        )
        """
    )
    try:
        conn.execute("ALTER TABLE v_prof_cache ADD COLUMN parts_json TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        conn.execute("ALTER TABLE v_prof_cache ADD COLUMN parts_schema INTEGER DEFAULT 1")
    except sqlite3.OperationalError:
        pass
    for stmt in (
        "ALTER TABLE cache_b_program_title ADD COLUMN hours_value REAL",
        "ALTER TABLE cache_pp_siz_table_header ADD COLUMN pp_hours REAL",
        "ALTER TABLE cache_pp_siz_table_header ADD COLUMN siz_hours REAL",
        "ALTER TABLE cache_pp_siz_table_header ADD COLUMN v_prof_hours REAL",
        "ALTER TABLE cache_pp_siz_table_header ADD COLUMN cache_schema INTEGER DEFAULT 1",
        "ALTER TABLE cache_v_registry ADD COLUMN col_d_hours REAL",
    ):
        try:
            conn.execute(stmt)
        except sqlite3.OperationalError:
            pass


def _excel_stat(path: Path) -> tuple[float, int] | None:
    if not path.is_file():
        return None
    st = path.stat()
    return (st.st_mtime, st.st_size)


def _sync_row(conn: sqlite3.Connection, role: str) -> tuple[str, float, int] | None:
    r = conn.execute(
        "SELECT path, mtime, size_bytes FROM excel_sync_state WHERE file_role = ?",
        (role,),
    ).fetchone()
    if not r:
        return None
    return (str(r[0]), float(r[1]), int(r[2]))


def _file_matches_sync(path: Path, role: str) -> bool:
    stat = _excel_stat(path)
    if stat is None:
        return False
    mtime, size = stat
    pr = str(path.resolve())
    with sqlite3.connect(_db_path()) as conn:
        ensure_excel_cache_tables(conn)
        row = _sync_row(conn, role)
    if row is None:
        return False
    p2, m2, s2 = row
    return (
        p2 == pr
        and s2 == size
        and abs(m2 - mtime) <= _MTIME_EPS
    )


def employees_excel_unchanged_since_cache(path: Path) -> bool:
    """True, если файл совпадает с последней синхронизацией роли employees."""
    return _file_matches_sync(path, EXCEL_SYNC_ROLE_EMPLOYEES)


def try_load_employees_from_cache(path: Path) -> list[EmployeeRecord] | None:
    """Сотрудников из БД не подгружаем — только из Excel (без хранения ПДн в snapshot)."""
    return None


def save_employees_cache(path: Path, records: list[EmployeeRecord]) -> None:
    """Удаляет устаревший снимок сотрудников для файла; новый снимок не создаётся."""
    del records  # загрузка из Excel уже в памяти
    try:
        pr = str(Path(path).resolve())
    except OSError:
        return
    with sqlite3.connect(_db_path()) as conn:
        ensure_excel_cache_tables(conn)
        conn.execute(
            "DELETE FROM cache_employees_snapshot WHERE file_path = ?",
            (pr,),
        )
        conn.execute(
            "DELETE FROM excel_sync_state WHERE file_role = ? AND path = ?",
            (EXCEL_SYNC_ROLE_EMPLOYEES, pr),
        )
        conn.commit()


def _ws_by_name(wb: Any, sheet_name: str) -> Any | None:
    names = {n.lower(): n for n in wb.sheetnames}
    sk = sheet_name.lower()
    if sk not in names:
        return None
    return wb[names[sk]]


def _v_program_ws(wb: Any) -> Any | None:
    m = {n.lower(): n for n in wb.sheetnames}
    for key in (V_PROGRAM_REGISTRY_SHEET.lower(), *V_PROGRAM_SHEET_NAME_ALIASES):
        lk = key.lower()
        if lk in m:
            return wb[m[lk]]
    return None


def _read_b_title_from_wb(wb: Any) -> tuple[str, int | None, float | None]:
    ws = _ws_by_name(wb, B_PROGRAM_SHEET_NAME)
    if ws is None:
        return "", None, None
    max_r = min(ws.max_row or 200, 500)
    for r in range(2, max_r + 1):
        v = ws.cell(row=r, column=B_PROGRAM_TITLE_COL).value
        if v is not None and str(v).strip():
            hv = parse_training_hours_value(ws.cell(row=r, column=B_PROGRAM_HOURS_COL).value)
            return str(v).strip(), r, hv
    return "", None, None


def _read_sheet_table_title_row_from_wb(wb: Any, sheet_name: str) -> tuple[str, int | None, float | None]:
    """Первая строка с названием на листе PP/SIZ (столбец 2) и часы из столбца 3 той же строки."""
    ws = _ws_by_name(wb, sheet_name)
    if ws is None:
        return "", None, None
    max_r = min(ws.max_row or 200, 500)
    for r in range(2, max_r + 1):
        v = ws.cell(row=r, column=B_PROGRAM_TITLE_COL).value
        if v is not None and str(v).strip():
            hv = parse_training_hours_value(ws.cell(row=r, column=B_PROGRAM_HOURS_COL).value)
            return str(v).strip(), r, hv
    return "", None, None


def _read_sheet_table_title_from_wb(wb: Any, sheet_name: str) -> str:
    """Первая непустая ячейка столбца B_PROGRAM_TITLE_COL со 2-й строки (как название на листе B)."""
    t, _, _ = _read_sheet_table_title_row_from_wb(wb, sheet_name)
    return t


def _read_v_prof_first_in_column_from_wb(wb: Any, column_one_based: int) -> str:
    ws = _ws_by_name(wb, V_PROF_SHEET_NAME)
    if ws is None:
        return ""
    max_r = min(ws.max_row or 200, 500)
    for r in range(2, max_r + 1):
        v = ws.cell(row=r, column=column_one_based).value
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def _read_v_registry_rows_from_wb(wb: Any) -> list[tuple[str, str, str, float | None]]:
    ws = _v_program_ws(wb)
    if ws is None:
        return []
    out: list[tuple[str, str, str, float | None]] = []
    max_r = min(ws.max_row or 0, V_PROGRAM_SHEET_MAX_ROWS)
    for r in range(2, max_r + 1):
        bv = ws.cell(row=r, column=V_PROGRAM_MATCH_COL_B).value
        if bv is None or not str(bv).strip():
            continue
        b = str(bv).strip()
        av = ws.cell(row=r, column=V_PROGRAM_GOS_REGISTRY_ID_COL_A).value
        gos_id = str(av).strip() if av is not None else ""
        cv = ws.cell(row=r, column=V_PROGRAM_TITLE_COL_C).value
        c = str(cv).strip() if cv is not None else ""
        dv = ws.cell(row=r, column=V_PROGRAM_HOURS_COL_D).value
        d_hours = parse_training_hours_value(dv)
        out.append((gos_id, b, c, d_hours))
    return out


def _rebuild_program_catalog_cache(
    conn: sqlite3.Connection, path: Path, mtime: float, size: int, pr: str
) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        title, src_row, b_hours = _read_b_title_from_wb(wb)
        from programs_v_prof import v_prof_layout_for_path

        layout = v_prof_layout_for_path(path)
        pp = _read_v_prof_first_in_column_from_wb(wb, layout.col_pp_one_based)
        siz = _read_v_prof_first_in_column_from_wb(wb, layout.col_siz_one_based)
        pp_table_title, _, pp_hours = _read_sheet_table_title_row_from_wb(wb, PP_TABLE_SHEET_NAME)
        siz_table_title, _, siz_hours = _read_sheet_table_title_row_from_wb(wb, SIZ_TABLE_SHEET_NAME)
        vrows = _read_v_registry_rows_from_wb(wb)
    finally:
        wb.close()

    conn.execute("DELETE FROM cache_v_registry WHERE file_path = ?", (pr,))
    conn.execute("DELETE FROM cache_b_program_title WHERE file_path = ?", (pr,))
    conn.execute("DELETE FROM cache_v_prof_column_snapshot WHERE file_path = ?", (pr,))
    conn.execute("DELETE FROM cache_pp_siz_table_header WHERE file_path = ?", (pr,))
    conn.execute(
        """
        INSERT OR REPLACE INTO cache_pp_siz_table_header
        (file_path, file_mtime, pp_title, siz_title, pp_hours, siz_hours, v_prof_hours, cache_schema)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            pr,
            mtime,
            pp_table_title,
            siz_table_title,
            pp_hours,
            siz_hours,
            None,
            PROGRAM_CATALOG_CACHE_SCHEMA,
        ),
    )
    conn.execute(
        """
        INSERT OR REPLACE INTO cache_b_program_title
        (file_path, file_mtime, title_text, source_row, hours_value)
        VALUES (?, ?, ?, ?, ?)
        """,
        (pr, mtime, title, src_row, b_hours),
    )
    for i, (g, b, c, d_h) in enumerate(vrows):
        conn.execute(
            """
            INSERT INTO cache_v_registry
            (file_path, file_mtime, row_order, gos_id, col_b, col_c, col_d_hours)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (pr, mtime, i, g, b, c, d_h),
        )
    for key, val in ((ProgramKey.PP.value, pp), (ProgramKey.SIZ.value, siz)):
        conn.execute(
            """
            INSERT OR REPLACE INTO cache_v_prof_column_snapshot
            (file_path, file_mtime, column_key, first_nonempty)
            VALUES (?, ?, ?, ?)
            """,
            (pr, mtime, key, val),
        )
    conn.execute("DELETE FROM v_prof_cache WHERE file_path = ?", (pr,))
    now = _utc_now_iso()
    conn.execute(
        """
        INSERT OR REPLACE INTO excel_sync_state
        (file_role, path, mtime, size_bytes, synced_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (EXCEL_SYNC_ROLE_PROGRAMS, pr, mtime, size, now),
    )


def ensure_program_catalog_cache(path: Path) -> None:
    """При необходимости перечитывает Excel и обновляет кэш программ (B, V, PP/SИЗ, очистка v_prof_cache)."""
    stat = _excel_stat(path)
    if stat is None:
        return
    mtime, size = stat
    pr = str(path.resolve())
    with sqlite3.connect(_db_path()) as conn:
        ensure_excel_cache_tables(conn)
        row = _sync_row(conn, EXCEL_SYNC_ROLE_PROGRAMS)
        if row is not None:
            p2, m2, s2 = row
            if p2 == pr and s2 == size and abs(m2 - mtime) <= _MTIME_EPS:
                ensure_excel_cache_tables(conn)
                row_hdr = conn.execute(
                    "SELECT cache_schema FROM cache_pp_siz_table_header WHERE file_path = ?",
                    (pr,),
                ).fetchone()
                if row_hdr is not None and (row_hdr[0] or 0) >= PROGRAM_CATALOG_CACHE_SCHEMA:
                    return
        _rebuild_program_catalog_cache(conn, path, mtime, size, pr)
        conn.commit()


def get_cached_b_program_title(path: Path) -> str:
    if not path.is_file():
        return ""
    ensure_program_catalog_cache(path)
    pr = str(path.resolve())
    with sqlite3.connect(_db_path()) as conn:
        ensure_excel_cache_tables(conn)
        r = conn.execute(
            "SELECT title_text FROM cache_b_program_title WHERE file_path = ?",
            (pr,),
        ).fetchone()
    return (r[0] or "").strip() if r else ""


def get_cached_pp_table_title(path: Path) -> str:
    """Наименование блока PP в таблице протокола — с листа PP (столбец как у «Б»)."""
    if not path.is_file():
        return ""
    ensure_program_catalog_cache(path)
    pr = str(path.resolve())
    with sqlite3.connect(_db_path()) as conn:
        ensure_excel_cache_tables(conn)
        r = conn.execute(
            "SELECT pp_title FROM cache_pp_siz_table_header WHERE file_path = ?",
            (pr,),
        ).fetchone()
    return (r[0] or "").strip() if r else ""


def get_cached_siz_table_title(path: Path) -> str:
    """Наименование блока СИЗ в таблице протокола — с листа SIZ."""
    if not path.is_file():
        return ""
    ensure_program_catalog_cache(path)
    pr = str(path.resolve())
    with sqlite3.connect(_db_path()) as conn:
        ensure_excel_cache_tables(conn)
        r = conn.execute(
            "SELECT siz_title FROM cache_pp_siz_table_header WHERE file_path = ?",
            (pr,),
        ).fetchone()
    return (r[0] or "").strip() if r else ""


def get_training_hours_for_program_key(path: Path | None, key: str) -> float | None:
    """Часы для блока «Б» / ПП / СИЗ из кэша файла программ; для «В» не используется (часы по строкам листа V)."""
    k = (key or "").strip().upper()
    if k not in ("B", "PP", "SIZ"):
        return None
    if path is None or not path.is_file():
        return None
    ensure_program_catalog_cache(path)
    pr = str(path.resolve())
    with sqlite3.connect(_db_path()) as conn:
        ensure_excel_cache_tables(conn)
        if k == "B":
            r = conn.execute(
                "SELECT hours_value FROM cache_b_program_title WHERE file_path = ?",
                (pr,),
            ).fetchone()
            return float(r[0]) if r and r[0] is not None else None
        r_ps = conn.execute(
            "SELECT pp_hours, siz_hours FROM cache_pp_siz_table_header WHERE file_path = ?",
            (pr,),
        ).fetchone()
        if not r_ps:
            return None
        if k == "PP":
            return float(r_ps[0]) if r_ps[0] is not None else None
        return float(r_ps[1]) if r_ps[1] is not None else None


def format_training_hours_ru(value: float) -> str:
    """Вывод суммы часов для текста протокола (целые без дробной части, иначе запятая)."""
    if value != value or value < 0:
        return "0"
    if abs(value - round(value)) < 1e-6:
        return str(int(round(value)))
    s = f"{value:.2f}".rstrip("0").rstrip(".").replace(".", ",")
    return s if s else "0"


def get_cached_v_registry_rows(path: Path) -> list[tuple[str, str, str, float | None]]:
    """Четвёрки (ID гос. реестра, B, C, часы из столбца D) с листа V."""
    if not path.is_file():
        return []
    ensure_program_catalog_cache(path)
    pr = str(path.resolve())
    with sqlite3.connect(_db_path()) as conn:
        ensure_excel_cache_tables(conn)
        cur = conn.execute(
            """
            SELECT gos_id, col_b, col_c, col_d_hours FROM cache_v_registry
            WHERE file_path = ? ORDER BY row_order
            """,
            (pr,),
        )
        out: list[tuple[str, str, str, float | None]] = []
        for a, b, c, d in cur.fetchall():
            dh = float(d) if d is not None else None
            out.append((str(a), str(b), str(c), dh))
        return out


def get_cached_v_prof_column(path: Path, column_one_based: int) -> str:
    """Первая непустая ячейка столбца V_PROF; PP/SИЗ из снимка, остальные — прямое чтение Excel."""
    if not path.is_file():
        return ""
    ensure_program_catalog_cache(path)
    from programs_v_prof import v_prof_layout_for_path

    layout = v_prof_layout_for_path(path)
    if column_one_based == V_PROF_TITLE_COL_PP:
        column_one_based = layout.col_pp_one_based
        key = ProgramKey.PP.value
    elif column_one_based == V_PROF_TITLE_COL_SIZ:
        column_one_based = layout.col_siz_one_based
        key = ProgramKey.SIZ.value
    else:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return ""
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return _read_v_prof_first_in_column_from_wb(wb, column_one_based)
        finally:
            wb.close()
    pr = str(path.resolve())
    with sqlite3.connect(_db_path()) as conn:
        ensure_excel_cache_tables(conn)
        r = conn.execute(
            """
            SELECT first_nonempty FROM cache_v_prof_column_snapshot
            WHERE file_path = ? AND column_key = ?
            """,
            (pr, key),
        ).fetchone()
    return (r[0] or "").strip() if r else ""


def invalidate_program_catalog_cache_for_path(path: Path | None = None) -> None:
    """
    Сброс кэша справочника программ (листы B, V, V_PROF и т.д.) для принудительного перечитывания с диска.
    Если передан path — дополнительно удаляются записи кэша, привязанные к этому файлу.
    Состояние синхронизации programs в excel_sync_state сбрасывается всегда.
    """
    pr = str(path.resolve()) if path is not None and path.is_file() else None
    with sqlite3.connect(_db_path()) as conn:
        ensure_excel_cache_tables(conn)
        conn.execute(
            "DELETE FROM excel_sync_state WHERE file_role = ?",
            (EXCEL_SYNC_ROLE_PROGRAMS,),
        )
        if pr:
            conn.execute("DELETE FROM cache_b_program_title WHERE file_path = ?", (pr,))
            conn.execute("DELETE FROM cache_v_registry WHERE file_path = ?", (pr,))
            conn.execute(
                "DELETE FROM cache_v_prof_column_snapshot WHERE file_path = ?", (pr,)
            )
            conn.execute("DELETE FROM cache_pp_siz_table_header WHERE file_path = ?", (pr,))
            conn.execute("DELETE FROM v_prof_cache WHERE file_path = ?", (pr,))
        conn.commit()


def invalidate_employees_cache_for_path(path: Path) -> None:
    """Сброс снимка сотрудников (например, при смене файла вручную)."""
    pr = str(path.resolve())
    with sqlite3.connect(_db_path()) as conn:
        ensure_excel_cache_tables(conn)
        conn.execute("DELETE FROM cache_employees_snapshot WHERE file_path = ?", (pr,))
        conn.execute(
            "DELETE FROM excel_sync_state WHERE file_role = ? AND path = ?",
            (EXCEL_SYNC_ROLE_EMPLOYEES, pr),
        )
        conn.commit()
