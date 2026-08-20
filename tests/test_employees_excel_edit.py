# -*- coding: utf-8 -*-
"""Запись сотрудников в Data_base.xlsx: добавление, архив, восстановление."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest import mock

from _bootstrap import setup_main_project_paths

setup_main_project_paths()

from employees_io import (
    EmployeeRecord,
    add_employee_to_excel,
    archive_employees_in_excel,
    employee_rows_for_excel_add,
    load_archived_employee_entries_from_excel,
    load_archived_employees_from_excel,
    load_employees_from_excel,
    restore_archived_employee_entries,
    restore_employees_from_archive,
    write_template_data_base_workbook,
    _analyze_employee_worksheet,
    _collect_employee_rows_from_sheet,
    _excel_cell_str,
    _last_employee_data_row,
)


class TestEmployeesExcelEdit(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.path = Path(self._tmp.name) / "Data_base.xlsx"
        write_template_data_base_workbook(self.path)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def test_add_employee_appends_row(self) -> None:
        rec = EmployeeRecord(
            fio="Иванов Иван Иванович",
            subdivision="Цех 1",
            profession="Слесарь",
            snils="12345678901",
        )
        add_employee_to_excel(self.path, rec, backup=False)
        rows = load_employees_from_excel(self.path)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].fio, rec.fio)
        self.assertEqual(rows[0].profession, rec.profession)

    def test_edit_preserves_formulas_on_other_sheet(self) -> None:
        """Запись сотрудников не должна уничтожать формулы на других листах (data_only=False)."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        wb = load_workbook(self.path)
        ws_extra = wb.create_sheet("calc")
        ws_extra["A1"] = 10
        ws_extra["B1"] = "=A1*2"
        wb.save(self.path)
        wb.close()
        add_employee_to_excel(
            self.path,
            EmployeeRecord(fio="Тестов Тест Тестович", profession="Слесарь"),
            backup=False,
        )
        wb2 = load_workbook(self.path, data_only=False)
        self.assertEqual(wb2["calc"]["B1"].value, "=A1*2")
        wb2.close()

    def test_writable_path_never_office_cache(self) -> None:
        from employees_io import employees_workbook_writable_path

        cache = Path(self._tmp.name) / ".office_cache"
        cache.mkdir(parents=True, exist_ok=True)
        cached = cache / "Data_base.ods.123.xlsx"
        cached.write_bytes(b"x")
        with mock.patch(
            "bundle_integration.office_cache_dir", return_value=cache
        ):
            out = employees_workbook_writable_path(cached)
        self.assertNotIn(".office_cache", str(out).replace("\\", "/"))
        self.assertTrue(str(out).endswith("Data_base.xlsx"))

    def test_writable_path_ods_to_sibling_xlsx(self) -> None:
        from employees_io import employees_workbook_writable_path

        ods = Path(self._tmp.name) / "Data_base.ods"
        out = employees_workbook_writable_path(ods)
        self.assertEqual(out, ods.with_suffix(".xlsx"))

    def test_writable_path_never_bundle_dir(self) -> None:
        from employees_io import employees_workbook_writable_path

        bundle = Path(self._tmp.name) / "bundle"
        user = Path(self._tmp.name) / "user_data"
        bundle.mkdir(parents=True, exist_ok=True)
        user.mkdir(parents=True, exist_ok=True)
        bundled = bundle / "Data_base.xlsx"
        bundled.write_bytes(b"x")
        with (
            mock.patch("app_paths.application_bundle_dir", return_value=bundle),
            mock.patch("app_paths.application_user_dir", return_value=user),
        ):
            out = employees_workbook_writable_path(bundled)
        self.assertEqual(out, user / "Data_base.xlsx")

    def test_employee_rows_for_excel_add_splits_profession2(self) -> None:
        rec = EmployeeRecord(
            fio="A",
            profession="Слесарь",
            profession2="Электрик",
        )
        rows = employee_rows_for_excel_add(rec)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].profession, "Слесарь")
        self.assertEqual(rows[1].profession, "Электрик")

    def test_add_employee_with_profession2_creates_two_rows(self) -> None:
        rec = EmployeeRecord(
            fio="Сидоров Сидор Сидорович",
            subdivision="Цех 2",
            profession="Слесарь",
            profession2="Электрик",
            snils="111",
        )
        n = add_employee_to_excel(self.path, rec, backup=False)
        self.assertEqual(n, 2)
        rows = load_employees_from_excel(self.path)
        self.assertEqual(len(rows), 2)
        professions = sorted(r.profession for r in rows)
        self.assertEqual(professions, ["Слесарь", "Электрик"])
        self.assertTrue(all(r.fio == rec.fio for r in rows))
        self.assertTrue(all(not (r.profession2 or "").strip() for r in rows))

    def test_archive_and_restore_roundtrip(self) -> None:
        rec = EmployeeRecord(
            fio="Петров Пётр Петрович",
            subdivision="Участок",
            profession="Электрик",
        )
        add_employee_to_excel(self.path, rec, backup=False)
        n = archive_employees_in_excel(self.path, [rec], backup=False)
        self.assertEqual(n, 1)
        self.assertEqual(load_employees_from_excel(self.path), [])
        archived = load_archived_employees_from_excel(self.path)
        self.assertEqual(len(archived), 1)
        self.assertEqual(archived[0].fio, rec.fio)
        r = restore_employees_from_archive(self.path, [rec], backup=False)
        self.assertEqual(r, 1)
        active = load_employees_from_excel(self.path)
        self.assertEqual(len(active), 1)
        self.assertEqual(active[0].fio, rec.fio)
        self.assertEqual(load_archived_employees_from_excel(self.path), [])

    def test_restore_by_fio_and_profession_only(self) -> None:
        """Восстановление по ФИО и должности из списка архива (без привязки к № строки)."""
        rec = EmployeeRecord(
            fio="Козлов Козел Козлович",
            subdivision="Цех 3",
            profession="Токарь",
        )
        add_employee_to_excel(self.path, rec, backup=False)
        archive_employees_in_excel(self.path, [rec], backup=False)
        archived = load_archived_employees_from_excel(self.path)
        target = EmployeeRecord(fio=rec.fio, profession=rec.profession, subdivision="")
        r = restore_employees_from_archive(self.path, [target], backup=False)
        self.assertEqual(r, 1)
        self.assertEqual(len(load_employees_from_excel(self.path)), 1)

    def test_restore_using_records_from_load_archived(self) -> None:
        """Как в UI «Архив…»: восстановление записей, прочитанных с листа архива."""
        rec = EmployeeRecord(
            fio="Козлов Козел Козлович",
            subdivision="Цех 3",
            profession="Токарь",
        )
        add_employee_to_excel(self.path, rec, backup=False)
        archive_employees_in_excel(self.path, [rec], backup=False)
        archived = load_archived_employees_from_excel(self.path)
        self.assertEqual(len(archived), 1)
        r = restore_employees_from_archive(self.path, archived, backup=False)
        self.assertEqual(r, 1)
        active = load_employees_from_excel(self.path)
        self.assertEqual(len(active), 1)
        self.assertEqual(active[0].profession, "Токарь")

    def test_restore_archive_match_by_fio_profession(self) -> None:
        """Восстановление по ФИО+должность, даже если profession2 в Excel заполнен."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        rec = EmployeeRecord(
            fio="Новиков Новик Новикович",
            subdivision="Участок 5",
            profession="Сварщик",
        )
        add_employee_to_excel(self.path, rec, backup=False)
        wb = load_workbook(self.path)
        ws = wb["rabotnik"]
        ws.cell(row=1, column=7, value="Совмещаемая профессия")
        ws.cell(row=2, column=7, value="лишнее совмещение")
        wb.save(self.path)
        wb.close()
        active = load_employees_from_excel(self.path)
        self.assertEqual(len(active), 1)
        self.assertEqual((active[0].profession2 or "").strip(), "лишнее совмещение")
        archive_employees_in_excel(self.path, active, backup=False)
        archived = load_archived_employees_from_excel(self.path)
        self.assertEqual((archived[0].profession2 or "").strip(), "лишнее совмещение")
        target = EmployeeRecord(
            fio=rec.fio,
            subdivision=rec.subdivision,
            profession=rec.profession,
            profession2="",
        )
        r = restore_employees_from_archive(self.path, [target], backup=False)
        self.assertEqual(r, 1)

    def test_restore_two_professions_same_fio_tab_header(self) -> None:
        """Data_base с Таб.№ и «Фамилия, И.»: два должности у одного ФИО — архив и восстановление."""
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        wb = Workbook()
        ws = wb.active
        ws.title = "rabotnik"
        ws.append(
            [
                "№ п/п",
                "Таб.№",
                "Фамилия, И.",
                "Подразделение",
                "Должность",
                "№ страхового свидетельства",
            ]
        )
        ws.append([1, 13000000, "Иванов Иван Иванович", "СНТ Ромашково", "Стропальщик", "123"])
        ws.append([1, 13000000, "Иванов Иван Иванович", "СНТ Ромашково", "Слесарь", "123"])
        wb.save(self.path)
        wb.close()
        rows = load_employees_from_excel(self.path)
        self.assertEqual(len(rows), 2)
        self.assertEqual({r.profession for r in rows}, {"Стропальщик", "Слесарь"})
        to_archive = [r for r in rows if r.profession == "Слесарь"]
        archive_employees_in_excel(self.path, to_archive, backup=False)
        wb_arch = load_workbook(self.path)
        ws_arch = wb_arch["rabotnik_archive"]
        arch_layout = _analyze_employee_worksheet(ws_arch)
        arch_serial = ws_arch.cell(row=2, column=arch_layout.serial_col + 1).value
        self.assertIn(arch_serial, (None, ""))
        wb_arch.close()
        archived = load_archived_employees_from_excel(self.path)
        self.assertEqual(len(archived), 1)
        self.assertEqual(archived[0].profession, "Слесарь")
        r = restore_employees_from_archive(self.path, archived, backup=False)
        self.assertEqual(r, 1)
        active = load_employees_from_excel(self.path)
        self.assertEqual(len(active), 2)
        professions = sorted(r.profession for r in active)
        self.assertEqual(professions, ["Слесарь", "Стропальщик"])
        wb2 = load_workbook(self.path)
        ws2 = wb2["rabotnik"]
        layout = _analyze_employee_worksheet(ws2)
        self.assertGreaterEqual(layout.cols.get("fio", -1), 0)
        last_row = _last_employee_data_row(ws2, layout)
        self.assertGreaterEqual(last_row, 3)
        restored_prof = ws2.cell(row=last_row, column=layout.col_prof + 1).value
        self.assertEqual(str(restored_prof).strip(), "Слесарь")
        if layout.serial_col >= 0:
            restored_serial = ws2.cell(row=last_row, column=layout.serial_col + 1).value
            self.assertIn(restored_serial, (None, ""))
        tab_val = ws2.cell(row=last_row, column=2).value
        self.assertEqual(tab_val, 13000000)
        wb2.close()

    def test_collect_employee_rows_finds_data_beyond_stale_max_row(self) -> None:
        """Поиск строк архива через iter_rows, а не ws.max_row (устаревший max_row)."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        rec = EmployeeRecord(fio="A B C", profession="X", subdivision="Y")
        add_employee_to_excel(self.path, rec, backup=False)
        archive_employees_in_excel(self.path, [rec], backup=False)
        wb = load_workbook(self.path)
        archive_ws = wb["rabotnik_archive"]
        layout = _analyze_employee_worksheet(archive_ws)
        rows_before = _collect_employee_rows_from_sheet(archive_ws, layout)
        self.assertEqual(len(rows_before), 1)
        wb.close()
        archived = load_archived_employees_from_excel(self.path)
        r = restore_employees_from_archive(self.path, archived, backup=False)
        self.assertEqual(r, 1)


class TestExcelCellStrSnils(unittest.TestCase):
    def test_int_like_float_without_dot_zero(self) -> None:
        self.assertEqual(_excel_cell_str((12345678901.0,), 0), "12345678901")
        self.assertEqual(_excel_cell_str((12345678901,), 0), "12345678901")
        self.assertNotIn(".0", _excel_cell_str((12345678901.0,), 0))


class TestBackupRotation(unittest.TestCase):
    def test_before_edit_keeps_rotated_copies(self) -> None:
        from employees_io import _backup_workbook_before_edit

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "Data_base.xlsx"
            path.write_text("v0", encoding="utf-8")
            for i in range(3):
                path.write_text(f"v{i + 1}", encoding="utf-8")
                _backup_workbook_before_edit(path, keep=3)
            primary = path.with_name("Data_base_before_edit.xlsx")
            b1 = path.with_name("Data_base_before_edit.1.xlsx")
            b2 = path.with_name("Data_base_before_edit.2.xlsx")
            self.assertTrue(primary.is_file())
            self.assertTrue(b1.is_file())
            self.assertTrue(b2.is_file())
            self.assertEqual(primary.read_text(encoding="utf-8"), "v3")
            self.assertEqual(b1.read_text(encoding="utf-8"), "v2")
            self.assertEqual(b2.read_text(encoding="utf-8"), "v1")


class TestArchiveMatchNorm(unittest.TestCase):
    def test_yo_and_spaces_match_for_archive(self) -> None:
        from employees_io import (
            EmployeeRecord,
            _employee_archive_records_match,
            employee_unique_key,
        )

        a = EmployeeRecord(fio="Алёна  Иванова", profession="Слесарь", subdivision="Цех 1")
        b = EmployeeRecord(fio="Алена Иванова", profession="Слесарь", subdivision="Цех 1")
        self.assertTrue(_employee_archive_records_match(a, b))
        self.assertEqual(employee_unique_key(a), employee_unique_key(b))


class TestRestoreByRowNum(unittest.TestCase):
    def test_restore_archived_entries_by_row_num(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "Data_base.xlsx"
            write_template_data_base_workbook(path)
            a = EmployeeRecord(fio="Первый П П", profession="Слесарь", subdivision="А")
            b = EmployeeRecord(fio="Второй В В", profession="Токарь", subdivision="Б")
            add_employee_to_excel(path, a, backup=False)
            add_employee_to_excel(path, b, backup=False)
            archive_employees_in_excel(path, [a, b], backup=False)
            entries = load_archived_employee_entries_from_excel(path)
            self.assertEqual(len(entries), 2)
            # Восстановить только одну строку по row_num
            one = [entries[0]]
            kept_fio = entries[1].record.fio
            n = restore_archived_employee_entries(path, one, backup=False)
            self.assertEqual(n, 1)
            left = load_archived_employee_entries_from_excel(path)
            self.assertEqual(len(left), 1)
            self.assertEqual(left[0].record.fio, kept_fio)
            active = load_employees_from_excel(path)
            self.assertEqual(len(active), 1)
            self.assertEqual(active[0].fio, one[0].record.fio)


if __name__ == "__main__":
    unittest.main()
